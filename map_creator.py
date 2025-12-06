import heapq
import json
import sqlite3
from dataclasses import dataclass, asdict
from typing import List, Dict, Set, Tuple, Optional
from enum import Enum
import requests
from datetime import datetime
import math
import os
import shutil
import geopandas as gpd


def get_tula_districts_from_osm():
    try:
        gdf = gpd.read_file('tula_districts/tula_administrative_districts.geojson')
        districts = {}
        for idx, row in gdf.iterrows():
            name = row['name']
            geom = row.geometry
            if geom.type == 'Polygon':
                coords = list(geom.exterior.coords)
                districts[name] = coords
            elif geom.type == 'MultiPolygon':
                largest = max(geom, key=lambda p: p.area)
                coords = list(largest.exterior.coords)
                districts[name] = coords
        return districts
    except Exception as e:
        print(f"Error loading districts from GeoJSON: {e}")
        return {}


class MobilityType(Enum):
    """–¢–∏–ø—ã –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–∏–π –º–æ–±–∏–ª—å–Ω–æ—Å—Ç–∏"""
    WHEELCHAIR = "–∫–æ–ª—è—Å–æ—á–Ω–∏–∫"
    VISUALLY_IMPAIRED = "—Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–π"
    CANE = "–æ–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å"


class AccessibilityFeature(Enum):
    """–¢–∏–ø—ã –æ–±—ä–µ–∫—Ç–æ–≤ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏"""
    RAMP_FOLDING = "–ø–∞–Ω–¥—É—Å_–æ—Ç–∫–∏–¥–Ω–æ–π"
    RAMP_FIXED = "–ø–∞–Ω–¥—É—Å_—Å—Ç–∞—Ü–∏–æ–Ω–∞—Ä–Ω—ã–π"
    TACTILE_GUIDING = "—Ç–∞–∫—Ç–∏–ª—å–Ω–∞—è_–ø–ª–∏—Ç–∫–∞_–Ω–∞–ø—Ä–∞–≤–ª—è—é—â–∞—è"
    TACTILE_WARNING = "—Ç–∞–∫—Ç–∏–ª—å–Ω–∞—è_–ø–ª–∏—Ç–∫–∞_–ø—Ä–µ–¥—É–ø—Ä–µ–∂–¥–∞—é—â–∞—è"
    CURB_LOWERING = "–ø–æ–Ω–∏–∂–µ–Ω–∏–µ_–±–æ—Ä–¥—é—Ä–∞"
    AUDIO_TRAFFIC_LIGHT = "—Å–≤–µ—Ç–æ—Ñ–æ—Ä_–∑–≤—É–∫–æ–≤–æ–π"
    WIDE_DOOR = "—à–∏—Ä–æ–∫–∞—è_–¥–≤–µ—Ä—å"
    HELP_BUTTON = "–∫–Ω–æ–ø–∫–∞_–≤—ã–∑–æ–≤–∞"
    HANDRAILS = "–ø–æ—Ä—É—á–Ω–∏"
    ELEVATOR = "–ª–∏—Ñ—Ç"
    ACCESSIBLE_PARKING = "–¥–æ—Å—Ç—É–ø–Ω–∞—è_–ø–∞—Ä–∫–æ–≤–∫–∞"


# –ê–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–∏–≤–Ω—ã–µ —Ä–∞–π–æ–Ω—ã –¢—É–ª—ã —Å –∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã–º–∏ –Ω–µ –ø–µ—Ä–µ—Å–µ–∫–∞—é—â–∏–º–∏—Å—è –≥—Ä–∞–Ω–∏—Ü–∞–º–∏ (–ø–æ–ª–∏–≥–æ–Ω—ã –≤ —Ñ–æ—Ä–º–∞—Ç–µ [lon, lat])
TULA_DISTRICTS = {
    "–¶–µ–Ω—Ç—Ä–∞–ª—å–Ω—ã–π": {
        "name": "–¶–µ–Ω—Ç—Ä–∞–ª—å–Ω—ã–π —Ä–∞–π–æ–Ω",
        "polygon": [
            [37.600, 54.180], [37.610, 54.180], [37.620, 54.180], [37.625, 54.180], [37.625, 54.190], [37.625, 54.200], [37.615, 54.200], [37.605, 54.200], [37.600, 54.200], [37.600, 54.190]
        ],
        "center": [37.612, 54.190],
        "color": "#ff6b6b"
    },
    "–°–æ–≤–µ—Ç—Å–∫–∏–π": {
        "name": "–°–æ–≤–µ—Ç—Å–∫–∏–π —Ä–∞–π–æ–Ω",
        "polygon": [
            [37.600, 54.200], [37.610, 54.200], [37.620, 54.200], [37.630, 54.200], [37.640, 54.200], [37.640, 54.210], [37.640, 54.220], [37.630, 54.220], [37.620, 54.220], [37.610, 54.220], [37.600, 54.220], [37.600, 54.210]
        ],
        "center": [37.620, 54.210],
        "color": "#4ecdc4"
    },
    "–ü—Ä–∏–≤–æ–∫–∑–∞–ª—å–Ω—ã–π": {
        "name": "–ü—Ä–∏–≤–æ–∫–∑–∞–ª—å–Ω—ã–π —Ä–∞–π–æ–Ω",
        "polygon": [
            [37.625, 54.180], [37.635, 54.180], [37.645, 54.180], [37.650, 54.180], [37.650, 54.190], [37.650, 54.200], [37.640, 54.200], [37.630, 54.200], [37.625, 54.200], [37.625, 54.190]
        ],
        "center": [37.637, 54.190],
        "color": "#45b7d1"
    },
    "–ó–∞—Ä–µ—á–µ–Ω—Å–∫–∏–π": {
        "name": "–ó–∞—Ä–µ—á–µ–Ω—Å–∫–∏–π —Ä–∞–π–æ–Ω",
        "polygon": [
            [37.580, 54.170], [37.590, 54.170], [37.600, 54.170], [37.610, 54.170], [37.610, 54.175], [37.610, 54.180], [37.610, 54.185], [37.600, 54.185], [37.590, 54.185], [37.580, 54.185], [37.580, 54.180], [37.580, 54.175]
        ],
        "center": [37.595, 54.177],
        "color": "#f9ca24"
    },
    "–ü—Ä–æ–ª–µ—Ç–∞—Ä—Å–∫–∏–π": {
        "name": "–ü—Ä–æ–ª–µ—Ç–∞—Ä—Å–∫–∏–π —Ä–∞–π–æ–Ω",
        "polygon": [
            [37.580, 54.185], [37.590, 54.185], [37.600, 54.185], [37.610, 54.185], [37.610, 54.190], [37.610, 54.195], [37.610, 54.200], [37.600, 54.200], [37.590, 54.200], [37.580, 54.200], [37.580, 54.195], [37.580, 54.190]
        ],
        "center": [37.595, 54.192],
        "color": "#6c5ce7"
    }
}


# Try to update with real boundaries from OSM
osm_districts = get_tula_districts_from_osm()
if osm_districts:
    name_mapping = {
        '–¶–µ–Ω—Ç—Ä–∞–ª—å–Ω—ã–π —Ä–∞–π–æ–Ω': '–¶–µ–Ω—Ç—Ä–∞–ª—å–Ω—ã–π',
        '–°–æ–≤–µ—Ç—Å–∫–∏–π —Ä–∞–π–æ–Ω': '–°–æ–≤–µ—Ç—Å–∫–∏–π',
        '–ü—Ä–∏–≤–æ–∫–∑–∞–ª—å–Ω—ã–π —Ä–∞–π–æ–Ω': '–ü—Ä–∏–≤–æ–∫–∑–∞–ª—å–Ω—ã–π',
        '–ó–∞—Ä–µ—á–µ–Ω—Å–∫–∏–π —Ä–∞–π–æ–Ω': '–ó–∞—Ä–µ—á–µ–Ω—Å–∫–∏–π',
        '–ü—Ä–æ–ª–µ—Ç–∞—Ä—Å–∫–∏–π —Ä–∞–π–æ–Ω': '–ü—Ä–æ–ª–µ—Ç–∞—Ä—Å–∫–∏–π'
    }
    for osm_name, coords in osm_districts.items():
        key = name_mapping.get(osm_name)
        if key:
            TULA_DISTRICTS[key]['polygon'] = coords
            print(f"Updated {key} with OSM data")


def point_in_polygon(x, y, polygon):
    """–ü—Ä–æ–≤–µ—Ä–∫–∞, –Ω–∞—Ö–æ–¥–∏—Ç—Å—è –ª–∏ —Ç–æ—á–∫–∞ –≤–Ω—É—Ç—Ä–∏ –ø–æ–ª–∏–≥–æ–Ω–∞ (–∞–ª–≥–æ—Ä–∏—Ç–º ray casting)"""
    n = len(polygon)
    inside = False
    p1x, p1y = polygon[0]
    for i in range(1, n + 1):
        p2x, p2y = polygon[i % n]
        if y > min(p1y, p2y):
            if y <= max(p1y, p2y):
                if x <= max(p1x, p2x):
                    if p1y != p2y:
                        xinters = (y - p1y) * (p2x - p1x) / (p2y - p1y) + p1x
                    if p1x == p2x or x <= xinters:
                        inside = not inside
        p1x, p1y = p2x, p2y
    return inside


def get_district_for_point(lat: float, lon: float) -> str:
    """–û–ø—Ä–µ–¥–µ–ª—è–µ—Ç —Ä–∞–π–æ–Ω –ø–æ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç–∞–º —Å –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–µ–º –ø–æ–ª–∏–≥–æ–Ω–æ–≤"""
    for district, data in TULA_DISTRICTS.items():
        polygon = data["polygon"]
        if point_in_polygon(lon, lat, polygon):  # Note: lon, lat for the function
            return district
    return "–ù–µ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω"


def get_district_statistics(db_path: str = "db/accessibility.db"):
    """–ü–æ–ª—É—á–∞–µ—Ç —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ –ø–æ —Ä–∞–π–æ–Ω–∞–º"""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ –æ–±—ä–µ–∫—Ç—ã
    cursor.execute("SELECT feature_type, latitude, longitude FROM accessibility_objects")
    objects = cursor.fetchall()
    conn.close()

    stats = {}
    for district in TULA_DISTRICTS.keys():
        stats[district] = {
            "name": TULA_DISTRICTS[district]["name"],
            "total_objects": 0,
            "by_type": {},
            "by_mobility": {
                "–∫–æ–ª—è—Å–æ—á–Ω–∏–∫": 0,
                "—Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–π": 0,
                "–æ–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å": 0,
                "–¥—Ä—É–≥–∏–µ": 0
            },
            "center": TULA_DISTRICTS[district]["center"],
            "color": TULA_DISTRICTS[district]["color"]
        }

    # –ú–∞–ø–ø–∏–Ω–≥ —Ç–∏–ø–æ–≤ –æ–±—ä–µ–∫—Ç–æ–≤ –∫ —Ç–∏–ø–∞–º –º–æ–±–∏–ª—å–Ω–æ—Å—Ç–∏
    mobility_mapping = {
        "–∫–æ–ª—è—Å–æ—á–Ω–∏–∫": ["–ø–∞–Ω–¥—É—Å_—Å—Ç–∞—Ü–∏–æ–Ω–∞—Ä–Ω—ã–π", "–ø–∞–Ω–¥—É—Å_–æ—Ç–∫–∏–¥–Ω–æ–π", "–ª–∏—Ñ—Ç", "—à–∏—Ä–æ–∫–∞—è_–¥–≤–µ—Ä—å", "–¥–æ—Å—Ç—É–ø–Ω–∞—è_–ø–∞—Ä–∫–æ–≤–∫–∞"],
        "—Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–π": ["—Ç–∞–∫—Ç–∏–ª—å–Ω–∞—è_–ø–ª–∏—Ç–∫–∞_–Ω–∞–ø—Ä–∞–≤–ª—è—é—â–∞—è", "—Ç–∞–∫—Ç–∏–ª—å–Ω–∞—è_–ø–ª–∏—Ç–∫–∞_–ø—Ä–µ–¥—É–ø—Ä–µ–∂–¥–∞—é—â–∞—è", "—Å–≤–µ—Ç–æ—Ñ–æ—Ä_–∑–≤—É–∫–æ–≤–æ–π", "–∫–Ω–æ–ø–∫–∞_–≤—ã–∑–æ–≤–∞"],
        "–æ–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å": ["–ø–æ—Ä—É—á–Ω–∏", "–ø–æ–Ω–∏–∂–µ–Ω–∏–µ_–±–æ—Ä–¥—é—Ä–∞"]
    }

    for obj in objects:
        feature_type, lat, lon = obj
        district = get_district_for_point(lat, lon)
        if district in stats:
            stats[district]["total_objects"] += 1
            if feature_type not in stats[district]["by_type"]:
                stats[district]["by_type"][feature_type] = 0
            stats[district]["by_type"][feature_type] += 1

            # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ç–∏–ø –º–æ–±–∏–ª—å–Ω–æ—Å—Ç–∏
            matched = False
            for mobility, types in mobility_mapping.items():
                if feature_type in types:
                    stats[district]["by_mobility"][mobility] += 1
                    matched = True
                    break
            if not matched:
                stats[district]["by_mobility"]["–¥—Ä—É–≥–∏–µ"] += 1

    return stats


def export_district_stats_to_excel(stats, filename="district_stats.xlsx"):
    """–≠–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–µ—Ç —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É —Ä–∞–π–æ–Ω–æ–≤ –≤ Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "–°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ —Ä–∞–π–æ–Ω–∞–º"

    # –ó–∞–≥–æ–ª–æ–≤–∫–∏
    headers = ["–†–∞–π–æ–Ω", "–í—Å–µ–≥–æ –æ–±—ä–µ–∫—Ç–æ–≤", "–ö–æ–ª—è—Å–æ—á–Ω–∏–∫–∏", "–°–ª–∞–±–æ–≤–∏–¥—è—â–∏–µ", "–û–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å", "–î—Ä—É–≥–∏–µ"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # –î–∞–Ω–Ω—ã–µ
    row = 2
    for district, data in stats.items():
        ws.cell(row=row, column=1).value = data["name"]
        ws.cell(row=row, column=2).value = data["total_objects"]
        ws.cell(row=row, column=3).value = data["by_mobility"]["–∫–æ–ª—è—Å–æ—á–Ω–∏–∫"]
        ws.cell(row=row, column=4).value = data["by_mobility"]["—Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–π"]
        ws.cell(row=row, column=5).value = data["by_mobility"]["–æ–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å"]
        ws.cell(row=row, column=6).value = data["by_mobility"]["–¥—Ä—É–≥–∏–µ"]
        row += 1

    # –ê–≤—Ç–æ–ø–æ–¥–±–æ—Ä —à–∏—Ä–∏–Ω—ã
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # –î–∏–∞–≥—Ä–∞–º–º–∞
    chart = BarChart()
    chart.title = "–°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ –ø–æ —Ä–∞–π–æ–Ω–∞–º"
    chart.y_axis.title = "–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –æ–±—ä–µ–∫—Ç–æ–≤"
    chart.x_axis.title = "–†–∞–π–æ–Ω—ã"

    data = Reference(ws, min_col=2, min_row=1, max_col=6, max_row=row-1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "G2")

    wb.save(filename)
    return filename


@dataclass
class AccessibilityObject:
    """–û–±—ä–µ–∫—Ç –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ –Ω–∞ –º–∞—Ä—à—Ä—É—Ç–µ"""
    id: Optional[int]
    feature_type: str
    description: str
    latitude: float
    longitude: float
    address: str
    created_at: Optional[str] = None


@dataclass
class RouteSegment:
    """–°–µ–≥–º–µ–Ω—Ç –º–∞—Ä—à—Ä—É—Ç–∞"""
    start_lat: float
    start_lon: float
    end_lat: float
    end_lon: float
    distance: float
    description: str
    accessibility_objects: List[AccessibilityObject]
    difficulty: float


# ===================================================================
# 1. AccessibilityDatabase ‚Äî 60 —É–Ω–∏–∫–∞–ª—å–Ω—ã—Ö –æ–±—ä–µ–∫—Ç–æ–≤ –≤ –¢—É–ª–µ (–ø–æ 20 –Ω–∞ —Ç–∏–ø)
# ===================================================================
class AccessibilityDatabase:
    def __init__(self, db_path: str = "db/accessibility.db"):
        self.db_path = db_path
        self.init_database()
        self.add_tula_accessibility_all()  # ‚Üê 60 –æ–±—ä–µ–∫—Ç–æ–≤!

    def init_database(self):
        # Ensure the database directory exists
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("""CREATE TABLE IF NOT EXISTS accessibility_objects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            feature_type TEXT NOT NULL,
            description TEXT,
            latitude REAL NOT NULL,
            longitude REAL NOT NULL,
            address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cursor.execute("""CREATE TABLE IF NOT EXISTS user_submissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            feature_type TEXT NOT NULL,
            description TEXT,
            address TEXT NOT NULL,
            photo_path TEXT,
            latitude REAL,
            longitude REAL,
            submitted_by TEXT,
            status TEXT DEFAULT 'pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cursor.execute("""CREATE TABLE IF NOT EXISTS admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            must_change_password INTEGER DEFAULT 1
        )""")
        # Insert default admin if not exists
        cursor.execute("SELECT COUNT(*) FROM admins WHERE username = 'admin'")
        if cursor.fetchone()[0] == 0:
            cursor.execute("INSERT INTO admins (username, password, must_change_password) VALUES (?, ?, ?)",
                           ('admin', generate_password_hash('admin'), 1))
        conn.commit()
        conn.close()

    def add_object(self, obj: AccessibilityObject) -> int:
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("""INSERT INTO accessibility_objects
            (feature_type, description, latitude, longitude, address)
            VALUES (?, ?, ?, ?, ?)""",
            (obj.feature_type, obj.description, obj.latitude, obj.longitude, obj.address))
        obj_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return obj_id

    def add_user_submission(self, feature_type: str, description: str, address: str, photo_path: str, lat: Optional[float] = None, lon: Optional[float] = None, submitted_by: str = "anonymous"):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("""INSERT INTO user_submissions
            (feature_type, description, address, photo_path, latitude, longitude, submitted_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (feature_type, description, address, photo_path, lat, lon, submitted_by))
        conn.commit()
        conn.close()

    def get_pending_submissions(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM user_submissions WHERE status = 'pending'")
        rows = cursor.fetchall()
        conn.close()
        return rows

    def approve_submission(self, submission_id: int):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("UPDATE user_submissions SET status = 'approved' WHERE id = ?", (submission_id,))
        # Move to main table if coordinates are available
        cursor.execute("SELECT feature_type, description, latitude, longitude, address FROM user_submissions WHERE id = ?", (submission_id,))
        row = cursor.fetchone()
        if row and row[2] is not None and row[3] is not None:
            cursor.execute("""INSERT INTO accessibility_objects
                (feature_type, description, latitude, longitude, address)
                VALUES (?, ?, ?, ?, ?)""", row)
        conn.commit()
        conn.close()

    def add_tula_accessibility_all(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM accessibility_objects")
        conn.commit()

        # –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º –æ–±—ä–µ–∫—Ç—ã –¥–ª—è –∫–∞–∂–¥–æ–≥–æ —Ä–∞–π–æ–Ω–∞
        all_objects = []
        for district, data in TULA_DISTRICTS.items():
            polygon = data["polygon"]
            min_lat = min(p[1] for p in polygon)
            max_lat = max(p[1] for p in polygon)
            min_lon = min(p[0] for p in polygon)
            max_lon = max(p[0] for p in polygon)

            # 4 –æ–±—ä–µ–∫—Ç–∞ –∫–∞–∂–¥–æ–≥–æ —Ç–∏–ø–∞ –Ω–∞ —Ä–∞–π–æ–Ω
            for i in range(4):
                lat = min_lat + (max_lat - min_lat) * (i + 0.5) / 4
                lon = min_lon + (max_lon - min_lon) * (i + 0.5) / 4

                # –ö–æ–ª—è—Å–æ—á–Ω–∏–∫–∏
                all_objects.append(AccessibilityObject(None, "–ø–∞–Ω–¥—É—Å_—Å—Ç–∞—Ü–∏–æ–Ω–∞—Ä–Ω—ã–π", f"–ü–∞–Ω–¥—É—Å –≤ {data['name']}", lat, lon, f"{data['name']}, –æ–±—ä–µ–∫—Ç {i+1}"))
                all_objects.append(AccessibilityObject(None, "–ª–∏—Ñ—Ç", f"–õ–∏—Ñ—Ç –≤ {data['name']}", lat + 0.001, lon + 0.001, f"{data['name']}, –æ–±—ä–µ–∫—Ç {i+1}"))
                all_objects.append(AccessibilityObject(None, "—à–∏—Ä–æ–∫–∞—è_–¥–≤–µ—Ä—å", f"–®–∏—Ä–æ–∫–∞—è –¥–≤–µ—Ä—å –≤ {data['name']}", lat - 0.001, lon - 0.001, f"{data['name']}, –æ–±—ä–µ–∫—Ç {i+1}"))
                all_objects.append(AccessibilityObject(None, "–¥–æ—Å—Ç—É–ø–Ω–∞—è_–ø–∞—Ä–∫–æ–≤–∫–∞", f"–ü–∞—Ä–∫–æ–≤–∫–∞ –≤ {data['name']}", lat + 0.002, lon + 0.002, f"{data['name']}, –æ–±—ä–µ–∫—Ç {i+1}"))

                # –°–ª–∞–±–æ–≤–∏–¥—è—â–∏–µ
                all_objects.append(AccessibilityObject(None, "—Ç–∞–∫—Ç–∏–ª—å–Ω–∞—è_–ø–ª–∏—Ç–∫–∞_–Ω–∞–ø—Ä–∞–≤–ª—è—é—â–∞—è", f"–¢–∞–∫—Ç–∏–ª—å–Ω–∞—è –ø–ª–∏—Ç–∫–∞ –≤ {data['name']}", lat, lon + 0.001, f"{data['name']}, –æ–±—ä–µ–∫—Ç {i+1}"))
                all_objects.append(AccessibilityObject(None, "—Å–≤–µ—Ç–æ—Ñ–æ—Ä_–∑–≤—É–∫–æ–≤–æ–π", f"–ó–≤—É–∫–æ–≤–æ–π —Å–≤–µ—Ç–æ—Ñ–æ—Ä –≤ {data['name']}", lat + 0.001, lon, f"{data['name']}, –æ–±—ä–µ–∫—Ç {i+1}"))
                all_objects.append(AccessibilityObject(None, "—Ç–∞–∫—Ç–∏–ª—å–Ω–∞—è_–ø–ª–∏—Ç–∫–∞_–ø—Ä–µ–¥—É–ø—Ä–µ–∂–¥–∞—é—â–∞—è", f"–ü—Ä–µ–¥—É–ø—Ä–µ–∂–¥–∞—é—â–∞—è –ø–ª–∏—Ç–∫–∞ –≤ {data['name']}", lat - 0.001, lon, f"{data['name']}, –æ–±—ä–µ–∫—Ç {i+1}"))
                all_objects.append(AccessibilityObject(None, "–∫–Ω–æ–ø–∫–∞_–≤—ã–∑–æ–≤–∞", f"–ö–Ω–æ–ø–∫–∞ –≤—ã–∑–æ–≤–∞ –≤ {data['name']}", lat, lon - 0.001, f"{data['name']}, –æ–±—ä–µ–∫—Ç {i+1}"))

                # –û–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å
                all_objects.append(AccessibilityObject(None, "–ø–æ—Ä—É—á–Ω–∏", f"–ü–æ—Ä—É—á–Ω–∏ –≤ {data['name']}", lat + 0.001, lon - 0.001, f"{data['name']}, –æ–±—ä–µ–∫—Ç {i+1}"))
                all_objects.append(AccessibilityObject(None, "–ø–æ–Ω–∏–∂–µ–Ω–∏–µ_–±–æ—Ä–¥—é—Ä–∞", f"–ü–æ–Ω–∏–∂–µ–Ω–∏–µ –±–æ—Ä–¥—é—Ä–∞ –≤ {data['name']}", lat - 0.001, lon + 0.001, f"{data['name']}, –æ–±—ä–µ–∫—Ç {i+1}"))

        for obj in all_objects:
            self.add_object(obj)

        print(f"–£–°–ü–ï–®–ù–û: –¥–æ–±–∞–≤–ª–µ–Ω–æ {len(all_objects)} –æ–±—ä–µ–∫—Ç–æ–≤ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ –≤ –¢—É–ª–µ (–ø–æ 12 –Ω–∞ –∫–∞–∂–¥—ã–π —Ç–∏–ø –≤ –∫–∞–∂–¥–æ–º —Ä–∞–π–æ–Ω–µ)!")
        conn.close()


class OpenStreetMapAPI:
    def __init__(self):
        self.base_url = "https://nominatim.openstreetmap.org"
        # –ò—Å–ø–æ–ª—å–∑—É–µ–º –ù–ê–î–Å–ñ–ù–´–ô —Å–µ—Ä–≤–µ—Ä, –∫–æ—Ç–æ—Ä—ã–π –†–ï–ê–õ–¨–ù–û –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ—Ç foot –≤ 2025
        self.routing_url = "https://routing.openstreetmap.de/routed-foot"
        # –ê–ª—å—Ç–µ—Ä–Ω–∞—Ç–∏–≤–∞: https://graphhopper.com/api/1/route (–Ω–æ –Ω—É–∂–µ–Ω –∫–ª—é—á)
        self.headers = {
            "User-Agent": "AccessibleNavigationApp/1.0 (+https://github.com/yourname/accessible-nav)"
        }

    def geocode(self, address: str) -> Optional[Tuple[float, float]]:
        # Default to Tula if no city specified
        if not any(city in address.lower() for city in ['—Ç—É–ª–∞', 'moscow', '—Å–ø–±', '–µ–∫–∞—Ç–µ—Ä–∏–Ω–±—É—Ä–≥']):
            address += ", –¢—É–ª–∞"
        try:
            response = requests.get(
                f"{self.base_url}/search",
                params={"q": address, "format": "json", "limit": 1, "countrycodes": "ru"},
                headers=self.headers,
                timeout=10
            )
            response.raise_for_status()
            data = response.json()
            if data:
                return float(data[0]["lat"]), float(data[0]["lon"])
        except Exception as e:
            print(f"–ì–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏–µ –æ—à–∏–±–∫–∞: {e}")
        return None

    def reverse_geocode(self, lat: float, lon: float) -> Optional[str]:
        try:
            response = requests.get(
                f"{self.base_url}/reverse",
                params={"lat": lat, "lon": lon, "format": "json", "addressdetails": 1},
                headers=self.headers,
                timeout=10
            )
            response.raise_for_status()
            data = response.json()
            if data and 'display_name' in data:
                return data['display_name']
        except Exception as e:
            print(f"–û–±—Ä–∞—Ç–Ω–æ–µ –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏–µ –æ—à–∏–±–∫–∞: {e}")
        return None

    def get_route(self, start: Tuple[float, float], end: Tuple[float, float]):
        return self.get_route_multi([start, end])

    def get_route_multi(self, points: List[Tuple[float, float]]):
        try:
            # –≠–¢–û–¢ —Å–µ—Ä–≤–µ—Ä –†–ï–ê–õ–¨–ù–û –¥–∞—ë—Ç –ø–µ—à–∏–π –º–∞—Ä—à—Ä—É—Ç!
            coords_str = ";".join(f"{p[1]},{p[0]}" for p in points)
            url = f"{self.routing_url}/route/v1/foot/{coords_str}"
            params = {
                "overview": "full",
                "geometries": "geojson",
                "steps": "true"
            }
            response = requests.get(url, params=params, timeout=25)
            response.raise_for_status()
            data = response.json()

            if data.get("code") != "Ok":
                print("OSRM –æ—à–∏–±–∫–∞:", data)
                return None, None

            route = data["routes"][0]
            coords = route["geometry"]["coordinates"]
            route_coords = [(lat, lon) for lon, lat in coords]

            return route_coords, route

        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ —Ä–æ—É—Ç–∏–Ω–≥–∞ (–ø–µ—à–∏–π): {e}")
            if 'response' in locals():
                print("–°–µ—Ä–≤–µ—Ä –æ—Ç–≤–µ—Ç–∏–ª:", response.text[:500])
        return None, None


# ===================================================================
# AccessibleNavigationSystem ‚Äî –£–ú–ù–´–ô –º–∞—Ä—à—Ä—É—Ç: –∫–æ—Ä–æ—Ç–∫–∏–π + –ø—Ä–∏–æ—Ä–∏—Ç–µ—Ç –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏
# ===================================================================
class AccessibleNavigationSystem:
    def __init__(self, db_path: str = "db/accessibility.db"):
        self.db = AccessibilityDatabase(db_path)
        self.osm = OpenStreetMapAPI()
        # –ü—Ä–∏–æ—Ä–∏—Ç–µ—Ç—ã –¥–ª—è –∫–∞–∂–¥–æ–≥–æ —Ç–∏–ø–∞
        self.feature_priorities = {
            MobilityType.WHEELCHAIR: {
                "–ø–∞–Ω–¥—É—Å_—Å—Ç–∞—Ü–∏–æ–Ω–∞—Ä–Ω—ã–π": 10, "–ª–∏—Ñ—Ç": 10, "—à–∏—Ä–æ–∫–∞—è_–¥–≤–µ—Ä—å": 8,
                "–¥–æ—Å—Ç—É–ø–Ω–∞—è_–ø–∞—Ä–∫–æ–≤–∫–∞": 7, "–ø–∞–Ω–¥—É—Å_–æ—Ç–∫–∏–¥–Ω–æ–π": 9
            },
            MobilityType.VISUALLY_IMPAIRED: {
                "—Ç–∞–∫—Ç–∏–ª—å–Ω–∞—è_–ø–ª–∏—Ç–∫–∞_–Ω–∞–ø—Ä–∞–≤–ª—è—é—â–∞—è": 10, "—Å–≤–µ—Ç–æ—Ñ–æ—Ä_–∑–≤—É–∫–æ–≤–æ–π": 10,
                "—Ç–∞–∫—Ç–∏–ª—å–Ω–∞—è_–ø–ª–∏—Ç–∫–∞_–ø—Ä–µ–¥—É–ø—Ä–µ–∂–¥–∞—é—â–∞—è": 9, "–∫–Ω–æ–ø–∫–∞_–≤—ã–∑–æ–≤–∞": 8
            },
            MobilityType.CANE: {
                "–ø–æ—Ä—É—á–Ω–∏": 10, "–ø–æ–Ω–∏–∂–µ–Ω–∏–µ_–±–æ—Ä–¥—é—Ä–∞": 9
            }
        }

    def find_route(self, start_address: str, end_address: str,
                    mobility_type: MobilityType,
                    user_location: Optional[Tuple[float, float]] = None,
                    start_coords: Optional[Tuple[float, float]] = None,
                    end_coords: Optional[Tuple[float, float]] = None) -> Dict:

        # 1. –ì–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏–µ
        if start_coords:
            start_coords_tuple = start_coords
            start_addr = "–í—ã–±—Ä–∞–Ω–Ω–æ–µ –º–µ—Å—Ç–æ –Ω–∞ –∫–∞—Ä—Ç–µ"
        elif start_address.lower() == "—Ç–µ–∫—É—â–∏–π" and user_location:
            start_coords_tuple = user_location
            start_addr = "–¢–µ–∫—É—â–µ–µ –º–µ—Å—Ç–æ–ø–æ–ª–æ–∂–µ–Ω–∏–µ"
        else:
            start_coords_tuple = self.osm.geocode(start_address)
            if not start_coords_tuple:
                return {"error": "–ù–µ —É–¥–∞–ª–æ—Å—å –Ω–∞–π—Ç–∏ –Ω–∞—á–∞–ª—å–Ω—ã–π –∞–¥—Ä–µ—Å"}
            start_addr = start_address

        if end_coords:
            end_coords_tuple = end_coords
            end_addr = "–í—ã–±—Ä–∞–Ω–Ω–æ–µ –º–µ—Å—Ç–æ –Ω–∞ –∫–∞—Ä—Ç–µ"
        else:
            end_coords_tuple = self.osm.geocode(end_address)
            if not end_coords_tuple:
                return {"error": "–ù–µ —É–¥–∞–ª–æ—Å—å –Ω–∞–π—Ç–∏ –∫–æ–Ω–µ—á–Ω—ã–π –∞–¥—Ä–µ—Å"}
            end_addr = end_address

        # 2. –°–Ω–∞—á–∞–ª–∞ —Å—Ç—Ä–æ–∏–º –°–ê–ú–´–ô –ö–û–†–û–¢–ö–ò–ô –º–∞—Ä—à—Ä—É—Ç
        base_route_coords, base_data = self.osm.get_route(start_coords_tuple, end_coords_tuple)
        if not base_route_coords:
            return {"error": "–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ—Å—Ç—Ä–æ–∏—Ç—å –º–∞—Ä—à—Ä—É—Ç"}

        base_distance = base_data["distance"]
        base_duration = int(base_data["duration"] / 60)

        start_lat, start_lon = start_coords_tuple
        end_lat, end_lon = end_coords_tuple

        # 3. –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º –æ–±—ä–µ–∫—Ç—ã –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ –≤ –æ–∫—Ä–µ—Å—Ç–Ω–æ—Å—Ç—è—Ö –º–∞—Ä—à—Ä—É—Ç–∞ (500–º - 1–∫–º)
        unique_objects = self.generate_accessibility_objects(base_route_coords, mobility_type)

        # 4. –í—ã–±–∏—Ä–∞–µ–º –¥–æ 6 –ª—É—á—à–∏—Ö –æ–±—ä–µ–∫—Ç–æ–≤ (–ø–æ –ø—Ä–∏–æ—Ä–∏—Ç–µ—Ç—É + –±–ª–∏–∑–æ—Å—Ç–∏ + –ø–æ—Ä—è–¥–∫—É —Å–ª–µ–¥–æ–≤–∞–Ω–∏—è)
        priorities = self.feature_priorities.get(mobility_type, {})

        # Compute cumulative distances along the route
        cum_dist = [0.0]
        for i in range(1, len(base_route_coords)):
            prev = base_route_coords[i-1]
            curr = base_route_coords[i]
            dist = ((curr[0] - prev[0])**2 + (curr[1] - prev[1])**2)**0.5
            cum_dist.append(cum_dist[-1] + dist)

        def score_object(obj):
            lat, lon, ftype, desc, addr, dist = obj
            priority = priorities.get(ftype, 0)
            # –ë–æ–Ω—É—Å –∑–∞ –±–ª–∏–∑–æ—Å—Ç—å –∫ –Ω–∞—á–∞–ª—É/–∫–æ–Ω—Ü—É
            start_dist = ((lat - start_lat)**2 + (lon - start_lon)**2)**0.5
            end_dist = ((lat - end_lat)**2 + (lon - end_lon)**2)**0.5
            position_bonus = max(0, 0.001 - min(start_dist, end_dist)) * 100000  # –±–æ–Ω—É—Å –∑–∞ –±–ª–∏–∑–æ—Å—Ç—å –∫ –Ω–∞—á–∞–ª—É/–∫–æ–Ω—Ü—É
            distance_penalty = dist * 500000  # —à—Ç—Ä–∞—Ñ –∑–∞ —É–¥–∞–ª–µ–Ω–Ω–æ—Å—Ç—å –æ—Ç –º–∞—Ä—à—Ä—É—Ç–∞
            return priority * 1000 + position_bonus - distance_penalty

        unique_objects.sort(key=score_object, reverse=True)
        best_objects = unique_objects[:6]

        # Add cumulative distance to each object and sort by position along route
        for i, obj in enumerate(best_objects):
            obj = list(obj)  # convert tuple to list
            lat, lon = obj[0], obj[1]
            min_d = float('inf')
            closest_idx = 0
            for idx, (rlat, rlon) in enumerate(base_route_coords):
                d = ((lat - rlat)**2 + (lon - rlon)**2)**0.5
                if d < min_d:
                    min_d = d
                    closest_idx = idx
            obj_cum_dist = cum_dist[closest_idx]
            obj.append(obj_cum_dist)
            best_objects[i] = obj  # update the list

        best_objects.sort(key=lambda x: x[6])  # sort by cumulative distance

        # Filter to ensure minimum distance along route to avoid close waypoints
        min_route_distance = 0.009  # ~1km along route
        filtered_objects = []
        for obj in best_objects:
            if not filtered_objects or all(abs(obj[6] - o[6]) > min_route_distance for o in filtered_objects):
                filtered_objects.append(obj)
        best_objects = filtered_objects[:3]  # limit to 3 waypoints max

        # 5. –°—Ç—Ä–æ–∏–º —Ñ–∏–Ω–∞–ª—å–Ω—ã–π –º–∞—Ä—à—Ä—É—Ç: —Å—Ç–∞—Ä—Ç ‚Üí –ª—É—á—à–∏–µ –æ–±—ä–µ–∫—Ç—ã ‚Üí —Ñ–∏–Ω–∏—à
        waypoints = [start_coords_tuple] + [(obj[0], obj[1]) for obj in best_objects] + [end_coords_tuple]
        used_objects = []

        for obj in best_objects:
            lat, lon, ftype, desc, addr = obj[:5]
            used_objects.append({
                "feature_type": ftype,
                "description": desc,
                "address": addr,
                "latitude": lat,
                "longitude": lon
            })

        # –°—Ç—Ä–æ–∏–º –º–∞—Ä—à—Ä—É—Ç —á–µ—Ä–µ–∑ –≤—ã–±—Ä–∞–Ω–Ω—ã–µ –æ–±—ä–µ–∫—Ç—ã –æ–¥–Ω–∏–º –∑–∞–ø—Ä–æ—Å–æ–º
        final_route, full_data = self.osm.get_route_multi(waypoints)

        if not final_route or full_data["distance"] > base_distance * 1.5:
            # –ï—Å–ª–∏ –∫—Ä—é–∫ —Å–ª–∏—à–∫–æ–º –±–æ–ª—å—à–æ–π –∏–ª–∏ –æ—à–∏–±–∫–∞ ‚Äî –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –∫–æ—Ä–æ—Ç–∫–∏–π –º–∞—Ä—à—Ä—É—Ç
            final_route = base_route_coords
            total_distance = base_distance
            total_minutes = base_duration
            used_objects = []  # –Ω–æ –≤—Å—ë —Ä–∞–≤–Ω–æ –ø–æ–∫–∞–∑—ã–≤–∞–µ–º –Ω–∞–π–¥–µ–Ω–Ω—ã–µ –æ–±—ä–µ–∫—Ç—ã –≤ –æ–ø–∏—Å–∞–Ω–∏–∏
        else:
            total_distance = full_data["distance"]
            total_minutes = int(full_data["duration"] / 60)

        description = self.generate_detailed_description(
            start_addr, end_address, total_distance, total_minutes, used_objects, mobility_type
        )

        return {
            "success": True,
            "start": {"address": start_addr, "coords": start_coords_tuple},
            "end": {"address": end_addr, "coords": end_coords_tuple},
            "route_coords": final_route,
            "accessibility_objects": used_objects,
            "description": description,
            "total_distance": int(total_distance),
            "duration_minutes": total_minutes,
            "mobility_type": mobility_type.value
        }

    def get_pedestrian_points_near_route(self, base_route_coords):
        # Get bbox around route
        lats = [p[0] for p in base_route_coords]
        lons = [p[1] for p in base_route_coords]
        min_lat, max_lat = min(lats), max(lats)
        min_lon, max_lon = min(lons), max(lons)
        # Expand by 0.01 degrees ~1km
        min_lat -= 0.01
        max_lat += 0.01
        min_lon -= 0.01
        max_lon += 0.01
        # Overpass query for pedestrian ways
        query = f"""
        [out:json];
        way["highway"~"footway|pedestrian|path"]({min_lat},{min_lon},{max_lat},{max_lon});
        out geom;
        """
        url = "https://overpass-api.de/api/interpreter"
        try:
            response = requests.post(url, data=query, timeout=10)
            data = response.json()
            points = []
            for way in data['elements']:
                if 'geometry' in way:
                    for geom in way['geometry']:
                        points.append((geom['lat'], geom['lon']))
            import random
            sampled = random.sample(points, min(50, len(points)))
            return sampled
        except:
            # Fallback to random
            return []

    def generate_accessibility_objects(self, base_route_coords, mobility_type):
        import random
        import math
        objects = []
        features = {
            MobilityType.WHEELCHAIR: ["–ø–∞–Ω–¥—É—Å_—Å—Ç–∞—Ü–∏–æ–Ω–∞—Ä–Ω—ã–π", "–ø–∞–Ω–¥—É—Å_–æ—Ç–∫–∏–¥–Ω–æ–π"],
            MobilityType.VISUALLY_IMPAIRED: ["—Ç–∞–∫—Ç–∏–ª—å–Ω–∞—è_–ø–ª–∏—Ç–∫–∞_–Ω–∞–ø—Ä–∞–≤–ª—è—é—â–∞—è", "—Å–≤–µ—Ç–æ—Ñ–æ—Ä_–∑–≤—É–∫–æ–≤–æ–π", "—Ç–∞–∫—Ç–∏–ª—å–Ω–∞—è_–ø–ª–∏—Ç–∫–∞_–ø—Ä–µ–¥—É–ø—Ä–µ–∂–¥–∞—é—â–∞—è", "–∫–Ω–æ–ø–∫–∞_–≤—ã–∑–æ–≤–∞"],
            MobilityType.CANE: ["–ø–æ—Ä—É—á–Ω–∏", "–ø–æ–Ω–∏–∂–µ–Ω–∏–µ_–±–æ—Ä–¥—é—Ä–∞"]
        }.get(mobility_type, [])
        pedestrian_points = self.get_pedestrian_points_near_route(base_route_coords)
        # Filter points within 500m of base_route
        def dist_to_route(lat, lon):
            return min(((lat - rlat)**2 + (lon - rlon)**2)**0.5 for rlat, rlon in base_route_coords)
        filtered_points = [p for p in pedestrian_points if dist_to_route(p[0], p[1]) < 0.005]  # ~500m
        if not filtered_points:
            # Fallback
            for lat, lon in base_route_coords[::20]:
                for _ in range(2):
                    dist_deg = random.uniform(0.001, 0.005)  # 100-500m
                    angle = random.uniform(0, 2 * math.pi)
                    new_lat = lat + dist_deg * math.cos(angle)
                    new_lon = lon + dist_deg * math.sin(angle)
                    feature = random.choice(features)
                    description = f"Generated {feature.replace('_', ' ')}"
                    address = f"Near pedestrian route at {new_lat:.4f}, {new_lon:.4f}"
                    obj_dist = dist_to_route(new_lat, new_lon)
                    obj = (new_lat, new_lon, feature, description, address, obj_dist)
                    objects.append(obj)
        else:
            for lat, lon in filtered_points[:20]:  # limit
                feature = random.choice(features)
                description = f"Generated {feature.replace('_', ' ')} on pedestrian route"
                address = f"On pedestrian route at {lat:.4f}, {lon:.4f}"
                obj_dist = dist_to_route(lat, lon)
                obj = (lat, lon, feature, description, address, obj_dist)
                objects.append(obj)
        return objects

    def generate_detailed_description(self, start_addr: str, end_addr: str,
                                      distance_m: float, duration_min: int,
                                      objects: List[dict], mobility_type: MobilityType) -> str:
        extra = " (—Å —É—á—ë—Ç–æ–º –æ–±—ä–µ–∫—Ç–æ–≤ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏)" if objects else " (—Å–∞–º—ã–π –∫–æ—Ä–æ—Ç–∫–∏–π)"
        desc = f"""–£–ú–ù–´–ô –ú–ê–†–®–†–£–¢ –î–õ–Ø {mobility_type.value.upper()}{extra}
{'='*70}
–û—Ç: {start_addr}
–î–æ: {end_addr}

–î–ª–∏–Ω–∞: {int(distance_m)} –º | –í—Ä–µ–º—è –≤ –ø—É—Ç–∏: {duration_min} –º–∏–Ω

–ö–õ–Æ–ß–ï–í–´–ï –û–ë–™–ï–ö–¢–´ –î–û–°–¢–£–ü–ù–û–°–¢–ò –ù–ê –ú–ê–†–®–†–£–¢–ï:
{'='*70}
"""
        if not objects:
            desc += "‚Üí –ú–∞—Ä—à—Ä—É—Ç –æ–ø—Ç–∏–º–∞–ª–µ–Ω. –û–±—ä–µ–∫—Ç—ã –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ –ø–æ–±–ª–∏–∑–æ—Å—Ç–∏ –Ω–µ –æ–±–Ω–∞—Ä—É–∂–µ–Ω—ã.\n"
        else:
            for i, obj in enumerate(objects, 1):
                name = obj["feature_type"].replace('_', ' ').title()
                desc += f"{i}. {name}\n   {obj['description']}\n   {obj['address']}\n\n"
            desc += "‚Üí –ú–∞—Ä—à—Ä—É—Ç –ø—Ä–æ—Ö–æ–¥–∏—Ç —á–µ—Ä–µ–∑ —ç—Ç–∏ –æ–±—ä–µ–∫—Ç—ã –¥–ª—è –≤–∞—à–µ–π –±–µ–∑–æ–ø–∞—Å–Ω–æ—Å—Ç–∏ –∏ –∫–æ–º—Ñ–æ—Ä—Ç–∞!\n"

        desc += "\n–ë–µ–∑–æ–ø–∞—Å–Ω–æ–≥–æ –ø—É—Ç–∏! –í—ã –¥–µ–ª–∞–µ—Ç–µ –º–∏—Ä –¥–æ—Å—Ç—É–ø–Ω–µ–µ ‚ôø"
        return desc


# Flask –≤–µ–±-–ø—Ä–∏–ª–æ–∂–µ–Ω–∏–µ
try:
    from flask import Flask, render_template_string, request, jsonify, redirect, url_for, send_from_directory, session, flash
    from flask_cors import CORS
    from werkzeug.utils import secure_filename
    from werkzeug.security import generate_password_hash, check_password_hash
    import os
    from xml_parser import XMLDataParser

    app = Flask(__name__)
    CORS(app)
    app.secret_key = 'supersecretkey'
    app.config['UPLOAD_FOLDER'] = 'uploads'
    app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    nav_system = AccessibleNavigationSystem()

    # Load organizations and infrastructure
    parser = XMLDataParser()
    try:
        parser.parse_organizations_xml("xml/–§–∞–π–ª_—Å–æ—Ü–ø–æ–¥–¥–µ—Ä–∂–∫–∞_1.xml")
        print(f"Loaded {len(parser.social_organizations)} organizations")
    except FileNotFoundError:
        print("Organizations XML not found, proceeding without organizations")

    # Using generated test data instead of XML for demo
    # try:
    #     parser.parse_infrastructure_xml("xml/–§–∞–π–ª_—Å–æ—Ü–ø–æ–¥–¥–µ—Ä–∂–∫–∞_2.xml")
    #     parser.populate_database(nav_system.db.db_path)
    #     print(f"Loaded infrastructure data from XML")
    # except FileNotFoundError:
    #     print("Infrastructure XML not found, using default data")

    organizations = parser.social_organizations
    
    HTML_TEMPLATE = r"""
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>–î–æ—Å—Ç—É–ø–Ω–∞—è –Ω–∞–≤–∏–≥–∞—Ü–∏—è</title>
        <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
        <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body { 
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                padding: 20px;
            }
            .container {
                max-width: 1400px;
                margin: 0 auto;
                background: white;
                border-radius: 20px;
                box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                overflow: hidden;
                position: relative;
            }
            .header {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 30px;
                text-align: center;
            }
            .header h1 { font-size: 2.5em; margin-bottom: 10px; }
            .header p { font-size: 1.2em; opacity: 0.9; }
            .header .links { margin-top: 20px; }
            .header .links a { color: white; margin: 0 10px; text-decoration: none; }
            .accessibility-buttons { margin-top: 20px; }
            .btn-accessibility {
                background: rgba(255,255,255,0.2);
                border: 1px solid rgba(255,255,255,0.3);
                color: white;
                padding: 8px 12px;
                margin: 0 5px;
                border-radius: 5px;
                cursor: pointer;
                transition: background 0.3s;
            }
            .btn-accessibility:hover { background: rgba(255,255,255,0.3); }
            .btn-accessibility.active { background: rgba(255,255,255,0.5); }
            body.high-contrast { background: black; color: white; }
            body.high-contrast .container { background: #333; }
            body.large-font { font-size: 1.2em; }
            body.large-font h1 { font-size: 3em; }
            body.large-font .btn { font-size: 1.3em; }
                background: rgba(255,255,255,0.2);
                color: white;
                border: 1px solid white;
                padding: 8px 12px;
                border-radius: 6px;
                font-size: 0.9em;
                cursor: pointer;
                margin: 0 5px;
                transition: all 0.3s;
            }
            .btn-accessibility:hover {
                background: rgba(255,255,255,0.3);
            }
            .high-contrast {
                background: #000 !important;
                color: #fff !important;
            }
            .high-contrast .sidebar {
                background: #111 !important;
            }
            .high-contrast .form-group label {
                color: #fff !important;
            }
            .high-contrast input, .high-contrast select {
                background: #333 !important;
                color: #fff !important;
                border-color: #fff !important;
            }
            .high-contrast .btn {
                background: #fff !important;
                color: #000 !important;
            }
            .content {
                display: grid;
                grid-template-columns: 400px 1fr;
                gap: 0;
            }
            .buttons {
                display: flex;
                justify-content: space-around;
                padding: 20px;
            }
            .sidebar {
                padding: 30px;
                border-right: 2px solid #f0f0f0;
                background: #fafafa;
            }
            .form-group {
                margin-bottom: 20px;
            }
            .form-group label {
                display: block;
                margin-bottom: 8px;
                font-weight: 600;
                color: #333;
            }
            .form-group input, .form-group select {
                width: 100%;
                padding: 12px;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                font-size: 1em;
                transition: border-color 0.3s;
            }
            .form-group input:focus, .form-group select:focus {
                outline: none;
                border-color: #667eea;
            }
            .btn {
                width: auto;
                padding: 15px;
                border: none;
                border-radius: 8px;
                font-size: 1.1em;
                font-weight: 600;
                cursor: pointer;
                transition: all 0.3s;
                flex: 1;
                min-width: 200px;
            }
            .button-row {
                display: flex;
                gap: 10px;
                flex-wrap: wrap;
            }
            .btn-primary {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
            }
            .btn-primary:hover {
                transform: translateY(-2px);
                box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
            }
            .btn-secondary {
                background: #f0f0f0;
                color: #333;
            }
            .btn-secondary:hover {
                background: #e0e0e0;
            }
            .btn-voice {
                background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
                color: white;
            }
            #map {
                height: calc(100vh - 200px);
                min-height: 500px;
                position: relative;
            }
            .route-info {
                margin-top: 20px;
                padding: 20px;
                background: white;
                border-radius: 10px;
                border-left: 4px solid #667eea;
            }
            .route-info h3 {
                margin-bottom: 15px;
                color: #667eea;
            }
            .route-info pre {
                white-space: pre-wrap;
                font-family: 'Courier New', monospace;
                font-size: 0.9em;
                line-height: 1.6;
                max-height: 400px;
                overflow-y: auto;
            }
            .loading {
                display: none;
                text-align: center;
                padding: 20px;
                color: #667eea;
            }
            .loading.active { display: block; }
            .spinner {
                border: 4px solid #f3f3f3;
                border-top: 4px solid #667eea;
                border-radius: 50%;
                width: 40px;
                height: 40px;
                animation: spin 1s linear infinite;
                margin: 0 auto;
            }
            @keyframes spin {
                0% { transform: rotate(0deg); }
                100% { transform: rotate(360deg); }
            }
            .icon {
                display: inline-block;
                margin-right: 8px;
            }
            .geolocation-status {
                font-size: 0.9em;
                color: #666;
                margin-top: 10px;
            }
            @keyframes pulse {
                0% { transform: scale(1); }
                50% { transform: scale(1.05); }
                100% { transform: scale(1); }
            }
            .modal {
                display: none;
                position: fixed;
                z-index: 2000;
                left: 0;
                top: 0;
                width: 100%;
                height: 100%;
                background-color: rgba(0,0,0,0.9);
            }
            .modal-content {
                position: absolute;
                top: 50%;
                left: 50%;
                transform: translate(-50%, -50%);
                background: transparent;
                border: none;
                padding: 0;
                width: auto;
                max-width: none;
            }
            .close {
                position: absolute;
                top: 20px;
                right: 20px;
                color: white;
                font-size: 36px;
                font-weight: bold;
                cursor: pointer;
                z-index: 2001;
            }
            .close:hover {
                color: red;
            }
            #samovarVideo {
                width: 100%;
            }
            @media (max-width: 900px) {
                .content {
                    grid-template-columns: 1fr;
                    grid-template-rows: 1fr auto;
                }
                .sidebar {
                    border-right: none;
                    border-bottom: 2px solid #f0f0f0;
                    order: 2;
                }
                #map {
                    order: 1;
                    height: 50vh;
                }
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>‚ôø –î–æ—Å—Ç—É–ø–Ω–∞—è –Ω–∞–≤–∏–≥–∞—Ü–∏—è</h1>
                <p>–ü–µ—Ä—Å–æ–Ω–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ –º–∞—Ä—à—Ä—É—Ç—ã –¥–ª—è –ª—é–¥–µ–π —Å –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–Ω—ã–º–∏ –≤–æ–∑–º–æ–∂–Ω–æ—Å—Ç—è–º–∏</p>
                <div class="accessibility-buttons">
                    <button id="elementVoiceBtn" class="btn-accessibility">üîä –û–∑–≤—É—á–∏–≤–∞–Ω–∏–µ —ç–ª–µ–º–µ–Ω—Ç–æ–≤</button>
                    <button id="routeVoiceBtn" class="btn-accessibility" style="display:none;">üîä –û–∑–≤—É—á–∏—Ç—å –º–∞—Ä—à—Ä—É—Ç</button>
                    <button id="contrastBtn" class="btn-accessibility">üëì –†–µ–∂–∏–º –¥–ª—è —Å–ª–∞–±–æ–≤–∏–¥—è—â–∏—Ö</button>
                </div>
            </div>
            <div class="content">
                <div class="sidebar">
                    <form id="routeForm">
                        <div class="form-group">
                            <label for="startAddress">
                                <span class="icon">üìç</span>–û—Ç–∫—É–¥–∞
                            </label>
                            <input type="text" id="startAddress" placeholder="–í–≤–µ–¥–∏—Ç–µ –∞–¥—Ä–µ—Å –∏–ª–∏ '—Ç–µ–∫—É—â–∏–π'" required title="–í–≤–µ–¥–∏—Ç–µ –∞–¥—Ä–µ—Å –æ—Ç–ø—Ä–∞–≤–ª–µ–Ω–∏—è –∏–ª–∏ '—Ç–µ–∫—É—â–∏–π' –¥–ª—è –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è –≥–µ–æ–ª–æ–∫–∞—Ü–∏–∏">
                            <div class="geolocation-status" id="geoStatus"></div>
                        </div>
                        
                        <div class="form-group">
                            <label for="endAddress">
                                <span class="icon">üéØ</span>–ö—É–¥–∞
                            </label>
                            <input type="text" id="endAddress" list="destinations" placeholder="–í–≤–µ–¥–∏—Ç–µ –∞–¥—Ä–µ—Å –∏–ª–∏ –≤—ã–±–µ—Ä–∏—Ç–µ –æ—Ä–≥–∞–Ω–∏–∑–∞—Ü–∏—é" required title="–í–≤–µ–¥–∏—Ç–µ –∞–¥—Ä–µ—Å –Ω–∞–∑–Ω–∞—á–µ–Ω–∏—è –∏–ª–∏ –≤—ã–±–µ—Ä–∏—Ç–µ –æ—Ä–≥–∞–Ω–∏–∑–∞—Ü–∏—é –∏–∑ —Å–ø–∏—Å–∫–∞">
                            <datalist id="destinations"></datalist>
                        </div>
                        
                        <div class="form-group">
                            <label for="mobilityType">
                                <span class="icon">üë§</span>–¢–∏–ø –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–∏–π
                            </label>
                            <select id="mobilityType" required title="–í—ã–±–µ—Ä–∏—Ç–µ —Ç–∏–ø –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–∏–π –º–æ–±–∏–ª—å–Ω–æ—Å—Ç–∏">
                                <option value="–∫–æ–ª—è—Å–æ—á–Ω–∏–∫">‚ôø –ö–æ–ª—è—Å–æ—á–Ω–∏–∫</option>
                                <option value="—Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–π">üëì –°–ª–∞–±–æ–≤–∏–¥—è—â–∏–π</option>
                                <option value="–æ–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å">ü¶Ø –û–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å</option>
                            </select>
                        </div>

                        <div class="button-row">
                        <button type="submit" class="btn btn-primary">
                            <span class="icon">üó∫Ô∏è</span>–ü–æ—Å—Ç—Ä–æ–∏—Ç—å –º–∞—Ä—à—Ä—É—Ç
                        </button>

                        <button type="button" class="btn btn-secondary" id="useLocationBtn">
                            <span class="icon">üì±</span>–ò—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å –º–æ—é –≥–µ–æ–ª–æ–∫–∞—Ü–∏—é
                        </button>

                        <a href="/submit" class="btn btn-secondary">
                            <span class="icon">‚ûï</span>–î–æ–±–∞–≤–∏—Ç—å –æ–±—ä–µ–∫—Ç –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏
                        </a>

                        </div>
                    </form>
                    
                    <div class="loading" id="loading">
                        <div class="spinner"></div>
                        <p>–ü–æ—Å—Ç—Ä–æ–µ–Ω–∏–µ –º–∞—Ä—à—Ä—É—Ç–∞...</p>
                    </div>
                    
                    <div class="route-info" id="routeInfo" style="display:none;">
                        <h3>–ò–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è –æ –º–∞—Ä—à—Ä—É—Ç–µ</h3>
                        <pre id="routeDescription"></pre>
                    </div>
                </div>
                
                <div id="map"></div>
            </div>

            <div id="pngContainer" style="display: flex; justify-content: space-around; padding: 20px; background: #f8f9fa; border-radius: 10px; margin: 20px;">
                <button id="png1" class="icon-btn" style="width: 50px; height: 50px; border: none; background: white; box-shadow: 0 4px 12px rgba(0,0,0,0.3); cursor: pointer;"><img src="/music/alien.png" style="width:100%; height:100%; object-fit:cover;" alt="icon"></button>
                <button id="png2" class="icon-btn" style="width: 50px; height: 50px; border: none; background: white; box-shadow: 0 4px 12px rgba(0,0,0,0.3); cursor: pointer;"><img src="/music/alien.png" style="width:100%; height:100%; object-fit:cover;" alt="icon"></button>
                <button id="png3" class="icon-btn" style="width: 50px; height: 50px; border: none; background: white; box-shadow: 0 4px 12px rgba(0,0,0,0.3); cursor: pointer;"><img src="/music/alien.png" style="width:100%; height:100%; object-fit:cover;" alt="icon"></button>
            </div>

            <!-- Notification -->
            <div id="notification" style="display:none; position:fixed; bottom:20px; left:50%; transform:translateX(-50%); background:white; padding:20px; border:1px solid black; z-index:2000; border-radius:10px; box-shadow:0 0 10px rgba(0,0,0,0.5);"></div>

            <!-- Video Modal -->
            <div id="videoModal" class="modal">
                <div class="modal-content">
                    <span id="closeVideoModal" class="close">&times;</span>
                    <video id="samovarVideo" autoplay loop style="display: block; margin: 0 auto; width: 90vw; height: 90vh; object-fit: contain;">
                        <source src="/music/samovar.mp4" type="video/mp4">
                        Your browser does not support the video tag.
                    </video>
                </div>
            </div>

        </div>
        
        <link href='https://unpkg.com/maplibre-gl@3.6.2/dist/maplibre-gl.css' rel='stylesheet' />
        <script src='https://unpkg.com/maplibre-gl@3.6.2/dist/maplibre-gl.js'></script>
        <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
        <audio id="bgMusic" preload="auto"></audio>

        <script>
            // –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è MapLibre GL JS
            const map = new maplibregl.Map({
                container: 'map',
                style: 'https://basemaps.cartocdn.com/gl/voyager-gl-style/style.json',
                center: [37.6175, 54.1931],
                zoom: 12,
                pitch: 30,
                bearing: 0
            });

            map.addControl(new maplibregl.NavigationControl());
            map.addControl(new maplibregl.GeolocateControl({
                positionOptions: { enableHighAccuracy: true },
                trackUserLocation: true
            }));

            // PNG click handling like the accessibility buttons above
            let pngClicked = 0;
            const totalPng = 3;

            ['png1', 'png2', 'png3'].forEach(id => {
                document.getElementById(id).addEventListener('click', () => {
                    document.getElementById(id).style.display = 'none';
                    pngClicked++;
                    const remaining = totalPng - pngClicked;

                    if (remaining > 0) {
                        showNotification(`üéâ –û—Ç–ª–∏—á–Ω–æ! –ù–∞–π–¥–µ–Ω–æ ${pngClicked}/${totalPng}. –û—Å—Ç–∞–ª–æ—Å—å ${remaining} –∏–∫–æ–Ω–æ–∫!`, 3000);
                    } else {
                        // All icons found - show video and audio
                        showNotification('üéä –ü–æ–∑–¥—Ä–∞–≤–ª—è–µ–º! –í—ã –Ω–∞—à–ª–∏ –≤—Å–µ –∏–∫–æ–Ω–∫–∏! –°–µ–π—á–∞—Å –≤–∫–ª—é—á–∏—Ç—Å—è –≤–∏–¥–µ–æ —Å —Å–∞–º–æ–≤–∞—Ä–æ–º!', 4000);
                        setTimeout(() => {
                            showVideoAndAudio();
                        }, 2000);
                    }
                });
            });

            let routeLayer = null;
            let routeSource = null;
            let accessibilityMarkers = [];
            let startMarker = null;
            let endMarker = null;
            let userLocationMarker = null;
            let addressMarkers = [];

            // –ü–æ–ª–Ω–∞—è –æ—á–∏—Å—Ç–∫–∞ –∫–∞—Ä—Ç—ã
            function clearMapCompletely() {
                if (routeLayer && map.getLayer('route')) map.removeLayer('route');
                if (routeSource && map.getSource('route')) map.removeSource('route');
                routeLayer = routeSource = null;

                accessibilityMarkers.forEach(m => m.remove());
                accessibilityMarkers = [];

                addressMarkers.forEach(m => m.remove());
                addressMarkers = [];

                if (startMarker) startMarker.remove();
                if (endMarker) endMarker.remove();
                if (userLocationMarker) userLocationMarker.remove();
                startMarker = endMarker = userLocationMarker = null;
            }


            // –û—Ç–æ–±—Ä–∞–∂–µ–Ω–∏–µ –º–∞—Ä—à—Ä—É—Ç–∞
            function displayRoute(data) {
                clearMapCompletely();

                const coords = data.route_coords.map(c => [c[1], c[0]]);

                map.addSource('route', {
                    type: 'geojson',
                    data: { type: 'Feature', geometry: { type: 'LineString', coordinates: coords } }
                });
                map.addLayer({
                    id: 'route',
                    type: 'line',
                    source: 'route',
                    paint: { 'line-color': '#667eea', 'line-width': 7, 'line-opacity': 0.9 }
                });
                routeSource = 'route';
                routeLayer = 'route';

                // –°—Ç–∞—Ä—Ç –∏ —Ñ–∏–Ω–∏—à
                startMarker = new maplibregl.Marker({ color: '#4ade80' })
                    .setLngLat(coords[0])
                    .setPopup(new maplibregl.Popup().setHTML(`<b>–ù–∞—á–∞–ª–æ</b><br>${data.start.address}`))
                    .addTo(map);

                endMarker = new maplibregl.Marker({ color: '#f87171' })
                    .setLngLat(coords[coords.length - 1])
                    .setPopup(new maplibregl.Popup().setHTML(`<b>–§–∏–Ω–∏—à</b><br>${data.end.address}`))
                    .addTo(map);

                // –û–±—ä–µ–∫—Ç—ã –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏
                const colorMap = {
                    '–ø–∞–Ω–¥—É—Å_—Å—Ç–∞—Ü–∏–æ–Ω–∞—Ä–Ω—ã–π': '#3b82f6', '–ª–∏—Ñ—Ç': '#8b5cf6', '—à–∏—Ä–æ–∫–∞—è_–¥–≤–µ—Ä—å': '#ec4899',
                    '–¥–æ—Å—Ç—É–ø–Ω–∞—è_–ø–∞—Ä–∫–æ–≤–∫–∞': '#06b6d4', '—Ç–∞–∫—Ç–∏–ª—å–Ω–∞—è_–ø–ª–∏—Ç–∫–∞_–Ω–∞–ø—Ä–∞–≤–ª—è—é—â–∞—è': '#f97316',
                    '—Å–≤–µ—Ç–æ—Ñ–æ—Ä_–∑–≤—É–∫–æ–≤–æ–π': '#10b981', '–ø–æ—Ä—É—á–Ω–∏': '#a16207', '–ø–æ–Ω–∏–∂–µ–Ω–∏–µ_–±–æ—Ä–¥—é—Ä–∞': '#84cc16'
                };

                data.accessibility_objects.forEach(obj => {
                    const marker = new maplibregl.Marker({ color: colorMap[obj.feature_type] || '#6b7280' })
                        .setLngLat([obj.longitude, obj.latitude])
                        .setPopup(new maplibregl.Popup({ offset: 25 }).setHTML(`
                            <b>${obj.feature_type.replace(/_/g, ' ').replace(/\b\w/g, l => l.toUpperCase())}</b><br>
                            ${obj.description}<br><small><i>${obj.address}</i></small>
                        `))
                        .addTo(map);
                    accessibilityMarkers.push(marker);
                });

                const bounds = new maplibregl.LngLatBounds(coords[0], coords[0]);
                coords.forEach(c => bounds.extend(c));
                map.fitBounds(bounds, { padding: 100, duration: 2000 });
            }

            // === –ê–í–¢–û–î–û–ü–û–õ–ù–ï–ù–ò–ï –î–õ–Ø startAddress ===
            const startInput = document.getElementById('startAddress');
            const startSuggestions = document.createElement('div');
            startSuggestions.className = 'suggestions';
            startInput.parentNode.style.position = 'relative';
            startInput.parentNode.appendChild(startSuggestions);

            // === –ê–í–¢–û–î–û–ü–û–õ–ù–ï–ù–ò–ï –î–õ–Ø endAddress ===
            const endInput = document.getElementById('endAddress');
            const endSuggestions = document.createElement('div');
            endSuggestions.className = 'suggestions';
            endInput.parentNode.style.position = 'relative';
            endInput.parentNode.appendChild(endSuggestions);

            // –°—Ç–∏–ª–∏ –¥–ª—è –ø–æ–¥—Å–∫–∞–∑–æ–∫
            const style = document.createElement('style');
            style.textContent = `
                .suggestions {
                    position: absolute;
                    top: 100%;
                    left: 0;
                    right: 0;
                    background: white;
                    border: 1px solid #ccc;
                    max-height: 200px;
                    overflow-y: auto;
                    z-index: 1000;
                    display: none;
                    box-shadow: 0 4px 10px rgba(0,0,0,0.1);
                    border-radius: 8px;
                    margin-top: 4px;
                }
                .suggestions div {
                    padding: 12px;
                    cursor: pointer;
                    border-bottom: 1px solid #eee;
                }
                .suggestions div:hover, .suggestions div.active {
                    background: #667eea;
                    color: white;
                }
            `;
            document.head.appendChild(style);

            // –£–Ω–∏–≤–µ—Ä—Å–∞–ª—å–Ω–∞—è —Ñ—É–Ω–∫—Ü–∏—è –∞–≤—Ç–æ–¥–æ–ø–æ–ª–Ω–µ–Ω–∏—è
            async function showSuggestions(input, box) {
                const query = input.value.trim();
                if (query.length < 2) {
                    box.style.display = 'none';
                    return;
                }

                try {
                    const res = await fetch(`/api/suggest_address?q=${encodeURIComponent(query)}`);
                    const suggestions = await res.json();

                    box.innerHTML = '';
                    if (suggestions.length === 0) {
                        box.style.display = 'none';
                        return;
                    }

                    suggestions.forEach((s, i) => {
                        const div = document.createElement('div');
                        div.textContent = s;
                        div.onclick = () => {
                            input.value = s;
                            box.style.display = 'none';
                        };
                        div.onmouseover = () => {
                            box.querySelectorAll('div').forEach(d => d.classList.remove('active'));
                            div.classList.add('active');
                        };
                        box.appendChild(div);
                    });

                    box.style.display = 'block';
                } catch (err) {
                    box.style.display = 'none';
                }
            }

            // –û–±—Ä–∞–±–æ—Ç—á–∏–∫–∏
            startInput.addEventListener('input', () => showSuggestions(startInput, startSuggestions));
            endInput.addEventListener('input', () => showSuggestions(endInput, endSuggestions));

            // –°–∫—Ä—ã—Ç–∏–µ –ø—Ä–∏ –∫–ª–∏–∫–µ –≤–Ω–µ
            document.addEventListener('click', e => {
                if (!e.target.closest('#startAddress') && !e.target.closest('.suggestions')) {
                    startSuggestions.style.display = 'none';
                }
                if (!e.target.closest('#endAddress') && !e.target.closest('.suggestions')) {
                    endSuggestions.style.display = 'none';
                }
            });

            // –ì–µ–æ–ª–æ–∫–∞—Ü–∏—è
            document.getElementById('useLocationBtn').addEventListener('click', () => {
                navigator.geolocation.getCurrentPosition(async pos => {
                    const lat = pos.coords.latitude;
                    const lon = pos.coords.longitude;

                    if (userLocationMarker) userLocationMarker.remove();
                    userLocationMarker = new maplibregl.Marker({ color: '#3b82f6' })
                        .setLngLat([lon, lat])
                        .setPopup(new maplibregl.Popup().setHTML('<b>–í—ã –∑–¥–µ—Å—å</b>'))
                        .addTo(map);

                    document.getElementById('startAddress').value = 'current';
                    document.getElementById('geoStatus').innerHTML = `–ì–µ–æ–ª–æ–∫–∞—Ü–∏—è: ¬±${pos.coords.accuracy.toFixed(0)} –º`;
                    document.getElementById('geoStatus').style.color = 'green';
                    map.flyTo({ center: [lon, lat], zoom: 16 });
                }, () => {
                    document.getElementById('geoStatus').textContent = '–ì–µ–æ–ª–æ–∫–∞—Ü–∏—è –Ω–µ–¥–æ—Å—Ç—É–ø–Ω–∞';
                    document.getElementById('geoStatus').style.color = 'red';
                });
            });

            // –ü–æ—Å—Ç—Ä–æ–µ–Ω–∏–µ –º–∞—Ä—à—Ä—É—Ç–∞
            document.getElementById('routeForm').addEventListener('submit', async e => {
                e.preventDefault();
                clearMapCompletely();

                const payload = {
                    start_address: document.getElementById('startAddress').value,
                    end_address: document.getElementById('endAddress').value,
                    mobility_type: document.getElementById('mobilityType').value
                };

                document.getElementById('loading').classList.add('active');

                try {
                    const res = await fetch('/api/route', { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: JSON.stringify(payload) });
                    const data = await res.json();

                    if (data.success) {
                        currentRoute = data;
                        displayRoute(data);
                        document.getElementById('routeDescription').textContent = data.description;
                        document.getElementById('routeInfo').style.display = 'block';
                        document.getElementById('routeVoiceBtn').style.display = 'inline-block';
                    } else {
                        alert('–û—à–∏–±–∫–∞: ' + (data.error || '–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–∞—è –æ—à–∏–±–∫–∞'));
                        document.getElementById('routeVoiceBtn').style.display = 'none';
                    }
                } catch (err) {
                    alert('–°–µ—Ä–≤–µ—Ä –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω');
                } finally {
                    document.getElementById('loading').classList.remove('active');
                }
            });

            // –ö–ª–∏–∫ –Ω–∞ –∫–∞—Ä—Ç–µ –¥–ª—è –≤—ã–±–æ—Ä–∞ –∞–¥—Ä–µ—Å–∞
            let selectedInput = null;
            document.getElementById('startAddress').addEventListener('focus', () => selectedInput = 'start');
            document.getElementById('endAddress').addEventListener('focus', () => selectedInput = 'end');

            map.on('click', async (e) => {
                if (!selectedInput) return;
                const { lng, lat } = e.lngLat;
                try {
                    const res = await fetch(`/api/reverse_geocode?lat=${lat}&lon=${lng}`);
                    const data = await res.json();
                    if (data.address) {
                        document.getElementById(selectedInput === 'start' ? 'startAddress' : 'endAddress').value = data.address;
                        selectedInput = null;
                    }
                } catch (err) {
                    console.error('Reverse geocode failed:', err);
                }
            });

            // –î–æ—Å—Ç—É–ø–Ω–æ—Å—Ç—å
            let elementVoiceMode = false;
            let highContrast = false;
            let largeFont = false;

            document.getElementById('elementVoiceBtn').addEventListener('click', () => {
                elementVoiceMode = !elementVoiceMode;
                document.getElementById('elementVoiceBtn').textContent = elementVoiceMode ? 'üîä –í—ã–∫–ª—é—á–∏—Ç—å –æ–∑–≤—É—á–∏–≤–∞–Ω–∏–µ' : 'üîä –û–∑–≤—É—á–∏–≤–∞–Ω–∏–µ —ç–ª–µ–º–µ–Ω—Ç–æ–≤';
                document.getElementById('elementVoiceBtn').classList.toggle('active', elementVoiceMode);
            });

            document.getElementById('contrastBtn').addEventListener('click', () => {
                highContrast = !highContrast;
                document.body.classList.toggle('high-contrast', highContrast);
                document.getElementById('contrastBtn').classList.toggle('active', highContrast);
                if (highContrast) {
                    // –û–∑–≤—É—á–∏—Ç—å –≤—Å–µ —ç–ª–µ–º–µ–Ω—Ç—ã –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å–∞ –¥–ª—è —Å–ª–∞–±–æ–≤–∏–¥—è—â–∏—Ö
                    announceInterface();
                }
            });

            // –§—É–Ω–∫—Ü–∏—è –æ–∑–≤—É—á–∏–≤–∞–Ω–∏—è —Å –±–æ–ª–µ–µ —á–µ–ª–æ–≤–µ—á–µ—Å–∫–∏–º –≥–æ–ª–æ—Å–æ–º
            function speakText(text, callback = null) {
                if (!('speechSynthesis' in window)) return;
                // –£–±–∏—Ä–∞–µ–º —ç–º–æ–¥–∑–∏ –∏–∑ —Ç–µ–∫—Å—Ç–∞
                text = text.replace(/[\u{1F600}-\u{1F64F}]|[\u{1F300}-\u{1F5FF}]|[\u{1F680}-\u{1F6FF}]|[\u{1F1E0}-\u{1F1FF}]|[\u{2600}-\u{26FF}]|[\u{2700}-\u{27BF}]/gu, '');
                const utter = new SpeechSynthesisUtterance(text);
                utter.lang = 'ru-RU';
                utter.rate = 0.9;  // –ë–æ–ª–µ–µ –µ—Å—Ç–µ—Å—Ç–≤–µ–Ω–Ω–∞—è —Å–∫–æ—Ä–æ—Å—Ç—å
                utter.pitch = 0.9; // –ë–æ–ª–µ–µ –Ω–∏–∑–∫–∞—è –≤—ã—Å–æ—Ç–∞ –¥–ª—è –º—É–∂—Å–∫–æ–≥–æ –≥–æ–ª–æ—Å–∞
                utter.volume = 0.9;
                // –ü–æ–ø—ã—Ç–∫–∞ –≤—ã–±—Ä–∞—Ç—å –º—É–∂—Å–∫–æ–π —Ä—É—Å—Å–∫–∏–π –≥–æ–ª–æ—Å
                const voices = speechSynthesis.getVoices();
                const russianVoice = voices.find(v => v.lang.startsWith('ru') && (v.name.includes('Male') || v.name.includes('–º—É–∂—Å–∫–æ–π') || !v.name.includes('Female')));
                if (russianVoice) utter.voice = russianVoice;
                if (callback) utter.onend = callback;
                speechSynthesis.speak(utter);
            }

            // –û–∑–≤—É—á–∫–∞ –º–∞—Ä—à—Ä—É—Ç–∞
            function announceRoute() {
                if (!currentRoute) return;
                speechSynthesis.cancel();
                const texts = [
                    `–ú–∞—Ä—à—Ä—É—Ç –æ—Ç ${currentRoute.start.address} –¥–æ ${currentRoute.end.address}`,
                    `–û–±—â–∞—è –¥–ª–∏–Ω–∞: ${currentRoute.total_distance} –º–µ—Ç—Ä–æ–≤. –ü—Ä–∏–º–µ—Ä–Ω–æ–µ –≤—Ä–µ–º—è –≤ –ø—É—Ç–∏: ${currentRoute.duration_minutes} –º–∏–Ω—É—Ç`,
                    ...currentRoute.accessibility_objects.map(o => `${o.feature_type.replace(/_/g, ' ').replace(/\b\w/g, l => l.toUpperCase())}: ${o.description}`),
                    "–ü—Ä–∏—è—Ç–Ω–æ–≥–æ –∏ –±–µ–∑–æ–ø–∞—Å–Ω–æ–≥–æ –ø—É—Ç–∏!"
                ];
                let i = 0;
                const speakNext = () => {
                    if (i >= texts.length) return;
                    speakText(texts[i++], speakNext);
                };
                speakNext();
            }

            // –û–∑–≤—É—á–∫–∞ –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å–∞ –¥–ª—è —Å–ª–∞–±–æ–≤–∏–¥—è—â–∏—Ö
            function announceInterface() {
                speechSynthesis.cancel();
                const elements = [
                    "–î–æ—Å—Ç—É–ø–Ω–∞—è –Ω–∞–≤–∏–≥–∞—Ü–∏—è –¥–ª—è –ª—é–¥–µ–π —Å –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–Ω—ã–º–∏ –≤–æ–∑–º–æ–∂–Ω–æ—Å—Ç—è–º–∏",
                    "–ü–æ–ª–µ –æ—Ç–∫—É–¥–∞ - –≤–≤–µ–¥–∏—Ç–µ –Ω–∞—á–∞–ª—å–Ω—ã–π –∞–¥—Ä–µ—Å –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ –¥–ª—è –≤—ã–±–æ—Ä–∞ –Ω–∞ –∫–∞—Ä—Ç–µ",
                    "–ü–æ–ª–µ –∫—É–¥–∞ - –≤–≤–µ–¥–∏—Ç–µ –∫–æ–Ω–µ—á–Ω—ã–π –∞–¥—Ä–µ—Å –∏–ª–∏ –≤—ã–±–µ—Ä–∏—Ç–µ –æ—Ä–≥–∞–Ω–∏–∑–∞—Ü–∏—é",
                    "–í—ã–±–µ—Ä–∏—Ç–µ —Ç–∏–ø –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–∏–π –º–æ–±–∏–ª—å–Ω–æ—Å—Ç–∏: –∫–æ–ª—è—Å–æ—á–Ω–∏–∫, —Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–π, –∏–ª–∏ –æ–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å",
                    "–ö–Ω–æ–ø–∫–∞ –ü–æ—Å—Ç—Ä–æ–∏—Ç—å –º–∞—Ä—à—Ä—É—Ç",
                    "–ö–Ω–æ–ø–∫–∞ –ò—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å –º–æ—é –≥–µ–æ–ª–æ–∫–∞—Ü–∏—é",
                    "–ö–Ω–æ–ø–∫–∞ –î–æ–±–∞–≤–∏—Ç—å –æ–±—ä–µ–∫—Ç –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏",
                    "–ö–Ω–æ–ø–∫–∞ –û–∑–≤—É—á–∏—Ç—å –º–∞—Ä—à—Ä—É—Ç - –¥–æ—Å—Ç—É–ø–Ω–∞ –ø–æ—Å–ª–µ –ø–æ—Å—Ç—Ä–æ–µ–Ω–∏—è –º–∞—Ä—à—Ä—É—Ç–∞",
                    "–ö–∞—Ä—Ç–∞ - –∫–ª–∏–∫–Ω–∏—Ç–µ –¥–ª—è –≤—ã–±–æ—Ä–∞ –∞–¥—Ä–µ—Å–∞"
                ];
                let i = 0;
                const speakNext = () => {
                    if (i >= elements.length) return;
                    speakText(elements[i++], speakNext);
                };
                speakNext();
            }

            document.getElementById('routeVoiceBtn').addEventListener('click', () => {
                if (!currentRoute) return;
                announceRoute();
            });

            // –û–∑–≤—É—á–∏–≤–∞–Ω–∏–µ —ç–ª–µ–º–µ–Ω—Ç–æ–≤ –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å–∞ –ø—Ä–∏ –≤–∫–ª—é—á–µ–Ω–Ω–æ–º —Ä–µ–∂–∏–º–µ
            function announceElement(element, eventType) {
                if (!elementVoiceMode || !('speechSynthesis' in window)) return;
                let text = '';
                if (element.tagName === 'INPUT' || element.tagName === 'SELECT' || element.tagName === 'TEXTAREA') {
                    const label = element.previousElementSibling ? element.previousElementSibling.textContent.trim() : element.placeholder || element.getAttribute('title') || '–ü–æ–ª–µ –≤–≤–æ–¥–∞';
                    text = label;
                    if (eventType === 'change' && element.tagName === 'SELECT') {
                        const selected = element.options[element.selectedIndex].text;
                        text += '. –í—ã–±—Ä–∞–Ω–æ: ' + selected;
                    }
                } else if (element.tagName === 'BUTTON') {
                    text = element.textContent.replace(/[^\w\s–∞-—è—ë]/gi, '').trim() || element.getAttribute('title') || '–ö–Ω–æ–ø–∫–∞';
                }
                if (text) {
                    speechSynthesis.speak(new SpeechSynthesisUtterance(text));
                }
            }

            // –î–æ–±–∞–≤–ª—è–µ–º listeners –∫–æ –≤—Å–µ–º –∏–Ω—Ç–µ—Ä–∞–∫—Ç–∏–≤–Ω—ã–º —ç–ª–µ–º–µ–Ω—Ç–∞–º
            document.querySelectorAll('input, select, textarea, button').forEach(el => {
                el.addEventListener('focus', () => announceElement(el, 'focus'));
                if (el.tagName === 'SELECT') {
                    el.addEventListener('change', () => announceElement(el, 'change'));
                }
                if (el.tagName === 'BUTTON') {
                    el.addEventListener('click', () => announceElement(el, 'click'));
                }
            });

            // –ü–æ–∫–∞–∑–∞—Ç—å –∞–¥—Ä–µ—Å–∞ –¥–æ–º–æ–≤
            async function showAddresses() {
                const bounds = map.getBounds();
                const bbox = `${bounds.getSouth()},${bounds.getWest()},${bounds.getNorth()},${bounds.getEast()}`;
                const query = `[out:json];way["building"]["addr:housenumber"](${bbox});out center meta;`;
                try {
                    const response = await fetch('https://overpass-api.de/api/interpreter', {
                        method: 'POST',
                        body: query
                    });
                    const data = await response.json();
                    // Clear previous
                    addressMarkers.forEach(m => m.remove());
                    addressMarkers = [];
                    // Add new
                    data.elements.slice(0, 100).forEach(element => {  // limit to 100
                        const lat = element.center.lat;
                        const lon = element.center.lon;
                        const housenumber = element.tags['addr:housenumber'] || '';
                        const street = element.tags['addr:street'] || '';
                        const address = `${street} ${housenumber}`.trim();
                        if (address) {
                            const marker = new maplibregl.Marker({ color: '#888' })
                                .setLngLat([lon, lat])
                                .setPopup(new maplibregl.Popup().setHTML(`<b>–ê–¥—Ä–µ—Å:</b><br>${address}`))
                                .addTo(map);
                            addressMarkers.push(marker);
                        }
                    });
                } catch (err) {
                    console.error('Error fetching addresses:', err);
                }
            }

            document.getElementById('showAddressesBtn').addEventListener('click', showAddresses);

            map.on('load', () => {
                console.log("MapLibre –≥–æ—Ç–æ–≤–∞ ‚Äî –≤—Å—ë –∏–¥–µ–∞–ª—å–Ω–æ!");
            });

            function showNotification(message, duration) {
                const notification = document.getElementById('notification');
                notification.textContent = message;
                notification.style.display = 'block';
                notification.style.background = 'linear-gradient(135deg, #667eea, #764ba2)';
                notification.style.color = 'white';
                notification.style.border = 'none';
                notification.style.fontSize = '16px';
                notification.style.fontWeight = 'bold';
                setTimeout(() => {
                    notification.style.display = 'none';
                }, duration);
            }

            function showVideoAndAudio() {
                const modal = document.getElementById('videoModal');
                modal.style.display = 'block';
                const video = document.getElementById('samovarVideo');
                video.play();
                const audio = document.getElementById('bgMusic');
                audio.src = '/music/julija-chicherina-tu-lu-la.mp3';
                audio.play();

                // Attach close button event
                const closeBtn = document.getElementById('closeVideoModal');
                if (closeBtn) {
                    closeBtn.addEventListener('click', closeModal);
                }
            }

            // Enhanced modal close functionality
            const modal = document.getElementById('videoModal');
            function closeModal() {
                const modal = document.getElementById('videoModal');
                modal.style.display = 'none';
                const video = document.getElementById('samovarVideo');
                video.pause();
                video.currentTime = 0; // Reset video
                const audio = document.getElementById('bgMusic');
                audio.pause();
                audio.currentTime = 0; // Reset audio
            }
            window.onclick = function(event) {
                if (event.target == modal) {
                    closeModal();
                }
            };

            // ESC key to close modal
            document.addEventListener('keydown', function(event) {
                if (event.key === 'Escape' && modal.style.display === 'block') {
                    closeModal();
                }
            });
        </script>
    </body>
    </html>
    """

    @app.route('/')
    def index():
        return render_template_string(HTML_TEMPLATE)

    @app.route('/api/route', methods=['POST'])
    def api_route():
        data = request.json
        start_address = data.get('start_address', '').strip()
        end_address = data.get('end_address', '').strip()
        mobility_type_str = data.get('mobility_type', '–∫–æ–ª—è—Å–æ—á–Ω–∏–∫')
        user_location = data.get('user_location')

        # –ü—Ä–µ–æ–±—Ä–∞–∑—É–µ–º —Å—Ç—Ä–æ–∫—É –≤ Enum
        try:
            mobility_type = MobilityType(mobility_type_str)
        except ValueError:
            return jsonify({"error": "–ù–µ–≤–µ—Ä–Ω—ã–π —Ç–∏–ø –º–æ–±–∏–ª—å–Ω–æ—Å—Ç–∏"}), 400

        # –î–∞–Ω–Ω—ã–µ —É–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã –≤ __init__

        result = nav_system.find_route(
            start_address=start_address,
            end_address=end_address,
            mobility_type=mobility_type,
            user_location=(user_location['lat'], user_location['lon']) if user_location else None,
            start_coords=data.get('start_coords'),
            end_coords=data.get('end_coords')
        )

        return jsonify(result)

    @app.route('/api/organizations')
    def api_organizations():
        mobility_type = request.args.get('mobility_type')
        orgs = []
        for org in organizations[:100]:  # Show more organizations
            serves = _matches_disability(org.served_disability_categories, mobility_type) if mobility_type else True
            orgs.append({
                "name": org.name,
                "address": org.address,
                "categories": org.served_disability_categories,
                "serves_current_type": serves,
                "warning": "–ù–µ –æ–±—Å–ª—É–∂–∏–≤–∞–µ—Ç –≤—ã–±—Ä–∞–Ω–Ω—ã–π —Ç–∏–ø –∏–Ω–≤–∞–ª–∏–¥–Ω–æ—Å—Ç–∏" if mobility_type and not serves else ""
            })
        return jsonify(orgs)

    def _matches_disability(categories, mobility_type):
        """Check if organization serves the given disability type"""
        # Map mobility types to possible category names in the XML
        mapping = {
            "–∫–æ–ª—è—Å–æ—á–Ω–∏–∫": ["–∏–Ω–≤–∞–ª–∏–¥—ã-–∫–æ–ª—è—Å–æ—á–Ω–∏–∫–∏", "–∫–æ–ª—è—Å–æ—á–Ω–∏–∫–∏", "–æ–ø–æ—Ä–Ω–æ-–¥–≤–∏–≥–∞—Ç–µ–ª—å–Ω—ã–µ", "–¥–≤–∏–≥–∞—Ç–µ–ª—å–Ω—ã–µ"],
            "—Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–π": ["—Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–µ", "–∏–Ω–≤–∞–ª–∏–¥—ã –ø–æ –∑—Ä–µ–Ω–∏—é", "—Å–ª–µ–ø—ã–µ", "–∑—Ä–µ–Ω–∏—è"],
            "–æ–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å": ["–∏–Ω–≤–∞–ª–∏–¥—ã —Å –ø–æ—Ä–∞–∂–µ–Ω–∏–µ–º –æ–ø–æ—Ä–Ω–æ-–¥–≤–∏–≥–∞—Ç–µ–ª—å–Ω–æ–≥–æ –∞–ø–ø–∞—Ä–∞—Ç–∞", "—Ç—Ä–∞–≤–º—ã", "–ø–æ–∂–∏–ª—ã–µ", "–¥–≤–∏–≥–∞—Ç–µ–ª—å–Ω—ã–µ"]
        }
        target_categories = mapping.get(mobility_type, [])
        return any(any(cat.lower() in category.lower() for cat in target_categories) for category in categories)

    def clean_address(full_address):
        parts = full_address.split(', ')
        cleaned = []
        for part in parts:
            part = part.strip()
            if not part:
                continue
            # Skip postal codes (5+ digits)
            if part.isdigit() and len(part) >= 5:
                continue
            # Skip country
            if part.lower() in ['—Ä–æ—Å—Å–∏—è', 'russia']:
                continue
            # Skip regions if too long
            if len(part) > 20 and any(word in part.lower() for word in ['–æ–±–ª–∞—Å—Ç—å', '–∫—Ä–∞–π', '—Ä–µ—Å–ø—É–±–ª–∏–∫–∞']):
                continue
            cleaned.append(part)
            if len(cleaned) >= 3:
                break
        return ', '.join(cleaned)
        parts = full_address.split(', ')
        cleaned = []
        for part in parts:
            part = part.strip()
            if not part:
                continue
            # Skip postal codes (5+ digits)
            if part.isdigit() and len(part) >= 5:
                continue
            # Skip country
            if part.lower() in ['—Ä–æ—Å—Å–∏—è', 'russia']:
                continue
            # Skip regions if too long
            if len(part) > 20 and any(word in part.lower() for word in ['–æ–±–ª–∞—Å—Ç—å', '–∫—Ä–∞–π', '—Ä–µ—Å–ø—É–±–ª–∏–∫–∞']):
                continue
            cleaned.append(part)
            if len(cleaned) >= 3:
                break
        return ', '.join(cleaned)

    @app.route('/api/suggest_address')
    def api_suggest_address():
        query = request.args.get('q', '').strip().lower()
        if not query:
            return jsonify([])
        conn = sqlite3.connect(nav_system.db.db_path)
        cursor = conn.cursor()
        # Get from accessibility_objects
        cursor.execute("SELECT DISTINCT address FROM accessibility_objects WHERE LOWER(address) LIKE ? LIMIT 5", ('%' + query + '%',))
        db_addresses = [row[0] for row in cursor.fetchall()]
        # Get from user_submissions
        cursor.execute("SELECT DISTINCT address FROM user_submissions WHERE LOWER(address) LIKE ? LIMIT 5", ('%' + query + '%',))
        db_addresses.extend([row[0] for row in cursor.fetchall()])
        conn.close()
        # Remove duplicates
        unique_addresses = list(set(db_addresses))[:5]
        if len(unique_addresses) < 5:
            # Fallback to OSM
            try:
                response = requests.get(
                    f"{nav_system.osm.base_url}/search",
                    params={"q": request.args.get('q', ''), "format": "json", "limit": 5 - len(unique_addresses), "countrycodes": "ru"},
                    headers=nav_system.osm.headers,
                    timeout=5
                )
                response.raise_for_status()
                data = response.json()
                osm_addresses = [clean_address(item['display_name']) for item in data]
                unique_addresses.extend(osm_addresses)
            except Exception as e:
                print(f"Suggest error: {e}")
        return jsonify(unique_addresses[:5])

    @app.route('/api/reverse_geocode')
    def api_reverse_geocode():
        lat = request.args.get('lat')
        lon = request.args.get('lon')
        if not lat or not lon:
            return jsonify({"error": "Missing lat/lon"})
        try:
            response = requests.get(
                f"{nav_system.osm.base_url}/reverse",
                params={"lat": lat, "lon": lon, "format": "json", "zoom": 18, "addressdetails": 1},
                headers=nav_system.osm.headers,
                timeout=5
            )
            response.raise_for_status()
            data = response.json()
            address = clean_address(data.get('display_name', ''))
            return jsonify({"address": address})
        except Exception as e:
            print(f"Reverse geocode error: {e}")
            return jsonify({"error": "Reverse geocoding failed"})

    @app.route('/submit')
    def submit_page():
        return render_template_string("""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>–î–æ–±–∞–≤–∏—Ç—å –æ–±—ä–µ–∫—Ç –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏</title>
            <style>
                * { margin: 0; padding: 0; box-sizing: border-box; }
                body {
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    padding: 20px;
                }
                .container {
                    max-width: 800px;
                    margin: 0 auto;
                    background: white;
                    border-radius: 20px;
                    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                    overflow: hidden;
                }
                .header {
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    padding: 30px;
                    text-align: center;
                }
                .header h1 { font-size: 2.5em; margin-bottom: 10px; }
                .header p { font-size: 1.2em; opacity: 0.9; }
                .header .admin-links { margin-top: 20px; }
                .header .admin-links a { color: white; margin: 0 10px; text-decoration: none; }
                .header .admin-links a[href="/admin/districts"] { background: rgba(255,255,255,0.2); padding: 5px 10px; border-radius: 5px; }
                .header .admin-links a[href="/admin/districts"] { background: rgba(255,255,255,0.2); padding: 5px 10px; border-radius: 5px; }
                .content {
                    padding: 30px;
                }
                .form-group {
                    margin-bottom: 20px;
                }
                .form-group label {
                    display: block;
                    margin-bottom: 8px;
                    font-weight: 600;
                    color: #333;
                }
                .form-group input, .form-group select, .form-group textarea {
                    width: 100%;
                    padding: 12px;
                    border: 2px solid #e0e0e0;
                    border-radius: 8px;
                    font-size: 1em;
                    transition: border-color 0.3s;
                }
                .form-group input:focus, .form-group select:focus, .form-group textarea:focus {
                    outline: none;
                    border-color: #667eea;
                }
                #routeForm {
                    display: flex;
                    flex-direction: column;
                    gap: 15px;
                }
                .btn {
                    width: auto;
                    padding: 15px;
                    border: none;
                    border-radius: 8px;
                    font-size: 1em;
                    font-weight: 600;
                    cursor: pointer;
                    transition: all 0.3s;
                    white-space: nowrap;
                    overflow: hidden;
                    text-overflow: ellipsis;
                    flex: 1;
                }
                .button-row {
                    display: flex;
                    gap: 10px;
                    flex-wrap: wrap;
                }
                .btn-primary {
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                }
                .btn-primary:hover {
                    transform: translateY(-2px);
                    box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
                }
                .btn-secondary {
                    background: #f0f0f0;
                    color: #333;
                }
                .btn-secondary:hover {
                    background: #e0e0e0;
                }
                .accessibility-buttons { margin-top: 20px; }
                .btn-accessibility {
                    background: #667eea;
                    color: white;
                    border: 1px solid #667eea;
                    padding: 10px 15px;
                    border-radius: 6px;
                    font-size: 0.9em;
                    cursor: pointer;
                    margin: 0 5px;
                    transition: all 0.3s;
                }
                .btn-accessibility:hover {
                    background: #5a67d8;
                }
                .high-contrast {
                    background: #000 !important;
                    color: #fff !important;
                }
                .high-contrast .container {
                    background: #111 !important;
                }
                .high-contrast .form-group label {
                    color: #fff !important;
                }
                .high-contrast input, .high-contrast select, .high-contrast textarea {
                    background: #333 !important;
                    color: #fff !important;
                    border-color: #fff !important;
                }
                .high-contrast .btn {
                    background: #fff !important;
                    color: #000 !important;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>‚ôø –î–æ–±–∞–≤–∏—Ç—å –æ–±—ä–µ–∫—Ç –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏</h1>
                    <p>–ü–æ–º–æ–≥–∏—Ç–µ —Å–¥–µ–ª–∞—Ç—å –≥–æ—Ä–æ–¥ –¥–æ—Å—Ç—É–ø–Ω–µ–µ</p>
                    <div class="accessibility-buttons">
                        <button id="voiceBtn" class="btn-accessibility">üîä –ì–æ–ª–æ—Å–æ–≤–æ–µ —Å–æ–ø—Ä–æ–≤–æ–∂–¥–µ–Ω–∏–µ</button>
                        <button id="contrastBtn" class="btn-accessibility">üëì –†–µ–∂–∏–º –¥–ª—è —Å–ª–∞–±–æ–≤–∏–¥—è—â–∏—Ö</button>
                    </div>
                </div>
                <div class="content">
                    <form action="/api/submit" method="post" enctype="multipart/form-data">
                        <div class="form-group">
                            <label>–¢–∏–ø –æ–±—ä–µ–∫—Ç–∞:</label>
                            <select name="feature_type" required title="–í—ã–±–µ—Ä–∏—Ç–µ —Ç–∏–ø –æ–±—ä–µ–∫—Ç–∞ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏">
                                <option value="–ø–∞–Ω–¥—É—Å_—Å—Ç–∞—Ü–∏–æ–Ω–∞—Ä–Ω—ã–π">–ü–∞–Ω–¥—É—Å —Å—Ç–∞—Ü–∏–æ–Ω–∞—Ä–Ω—ã–π</option>
                                <option value="–ø–∞–Ω–¥—É—Å_–æ—Ç–∫–∏–¥–Ω–æ–π">–ü–∞–Ω–¥—É—Å –æ—Ç–∫–∏–¥–Ω–æ–π</option>
                                <option value="–ª–∏—Ñ—Ç">–õ–∏—Ñ—Ç</option>
                                <option value="—Ç–∞–∫—Ç–∏–ª—å–Ω–∞—è_–ø–ª–∏—Ç–∫–∞_–Ω–∞–ø—Ä–∞–≤–ª—è—é—â–∞—è">–¢–∞–∫—Ç–∏–ª—å–Ω–∞—è –ø–ª–∏—Ç–∫–∞ –Ω–∞–ø—Ä–∞–≤–ª—è—é—â–∞—è</option>
                                <option value="—Å–≤–µ—Ç–æ—Ñ–æ—Ä_–∑–≤—É–∫–æ–≤–æ–π">–ó–≤—É–∫–æ–≤–æ–π —Å–≤–µ—Ç–æ—Ñ–æ—Ä</option>
                                <option value="–ø–æ—Ä—É—á–Ω–∏">–ü–æ—Ä—É—á–Ω–∏</option>
                                <option value="–ø–æ–Ω–∏–∂–µ–Ω–∏–µ_–±–æ—Ä–¥—é—Ä–∞">–ü–æ–Ω–∏–∂–µ–Ω–∏–µ –±–æ—Ä–¥—é—Ä–∞</option>
                            </select>
                        </div>
                        <div class="form-group">
                            <label>–û–ø–∏—Å–∞–Ω–∏–µ:</label>
                            <textarea name="description" required title="–û–ø–∏—à–∏—Ç–µ –æ–±—ä–µ–∫—Ç –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ –ø–æ–¥—Ä–æ–±–Ω–æ"></textarea>
                        </div>
                        <div class="form-group">
                            <label>–ê–¥—Ä–µ—Å:</label>
                            <input type="text" name="address" required title="–í–≤–µ–¥–∏—Ç–µ –ø–æ–ª–Ω—ã–π –∞–¥—Ä–µ—Å –æ–±—ä–µ–∫—Ç–∞ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏">
                        </div>
                        <div class="form-group">
                            <label>–§–æ—Ç–æ:</label>
                            <input type="file" name="photo" accept="image/*" required title="–ó–∞–≥—Ä—É–∑–∏—Ç–µ —Ñ–æ—Ç–æ –æ–±—ä–µ–∫—Ç–∞ (–∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ)">
                        </div>
                        <div class="button-row">
                        <button type="submit" class="btn btn-primary">–û—Ç–ø—Ä–∞–≤–∏—Ç—å –Ω–∞ –ø—Ä–æ–≤–µ—Ä–∫—É</button>
                        <a href="/" class="btn btn-secondary">–ù–∞–∑–∞–¥</a>
                        </div>
                    </form>
                </div>
            </div>
        </body>
        <script>
            let submitSuggestionBox = document.createElement('div');
            submitSuggestionBox.id = 'submitSuggestions';
            submitSuggestionBox.style.cssText = `position: absolute; background: white; border: 1px solid #ccc; max-height: 200px; overflow-y: auto; z-index: 1000; display: none; width: 100%; box-shadow: 0 2px 4px rgba(0,0,0,0.1);`;
            document.querySelector('input[name="address"]').parentNode.style.position = 'relative';
            document.querySelector('input[name="address"]').parentNode.appendChild(submitSuggestionBox);

            let submitSelectedIndex = -1;

            function updateSubmitSelection() {
                const items = submitSuggestionBox.children;
                for (let i = 0; i < items.length; i++) {
                    items[i].style.background = i === submitSelectedIndex ? '#667eea' : 'white';
                    items[i].style.color = i === submitSelectedIndex ? 'white' : 'black';
                }
            }

            document.querySelector('input[name="address"]').addEventListener('input', async e => {
                const query = e.target.value;
                if (query.length < 1) {
                    submitSuggestionBox.style.display = 'none';
                    return;
                }
                try {
                    const res = await fetch('/api/suggest_address?q=' + encodeURIComponent(query));
                    const suggestions = await res.json();
                    submitSuggestionBox.innerHTML = '';
                    submitSelectedIndex = -1;
                    suggestions.forEach((s, index) => {
                        const div = document.createElement('div');
                        div.textContent = s;
                        div.style.cssText = 'padding: 8px; cursor: pointer; border-bottom: 1px solid #eee;';
                        div.addEventListener('click', () => {
                            e.target.value = s;
                            submitSuggestionBox.style.display = 'none';
                        });
                        div.addEventListener('mouseover', () => {
                            submitSelectedIndex = index;
                            updateSubmitSelection();
                        });
                        submitSuggestionBox.appendChild(div);
                    });
                    submitSuggestionBox.style.display = suggestions.length ? 'block' : 'none';
                } catch (err) {
                    submitSuggestionBox.style.display = 'none';
                }
            });

            document.querySelector('input[name="address"]').addEventListener('keydown', e => {
                const items = submitSuggestionBox.children;
                if (submitSuggestionBox.style.display === 'none' || items.length === 0) return;
                if (e.key === 'ArrowDown') {
                    e.preventDefault();
                    submitSelectedIndex = (submitSelectedIndex + 1) % items.length;
                    updateSubmitSelection();
                } else if (e.key === 'ArrowUp') {
                    e.preventDefault();
                    submitSelectedIndex = submitSelectedIndex <= 0 ? items.length - 1 : submitSelectedIndex - 1;
                    updateSubmitSelection();
                } else if (e.key === 'Enter') {
                    e.preventDefault();
                    if (submitSelectedIndex >= 0) {
                        items[submitSelectedIndex].click();
                    }
                } else if (e.key === 'Escape') {
                    submitSuggestionBox.style.display = 'none';
                    submitSelectedIndex = -1;
                }
            });

            document.addEventListener('click', (e) => {
                if (!e.target.closest('input[name="address"]') && !e.target.closest('#submitSuggestions')) {
                    submitSuggestionBox.style.display = 'none';
                    submitSelectedIndex = -1;
                }
            });

            // Accessibility features
            let voiceMode = false;
            let highContrast = false;

            document.getElementById('voiceBtn').addEventListener('click', () => {
                voiceMode = !voiceMode;
                document.getElementById('voiceBtn').textContent = voiceMode ? 'üîä –í—ã–∫–ª—é—á–∏—Ç—å –≥–æ–ª–æ—Å' : 'üîä –ì–æ–ª–æ—Å–æ–≤–æ–µ —Å–æ–ø—Ä–æ–≤–æ–∂–¥–µ–Ω–∏–µ';
            });

            document.getElementById('contrastBtn').addEventListener('click', () => {
                highContrast = !highContrast;
                document.body.classList.toggle('high-contrast', highContrast);
                document.getElementById('contrastBtn').textContent = highContrast ? 'üëì –û–±—ã—á–Ω—ã–π —Ä–µ–∂–∏–º' : 'üëì –†–µ–∂–∏–º –¥–ª—è —Å–ª–∞–±–æ–≤–∏–¥—è—â–∏—Ö';
            });

            // Voice announcements for inputs, selects, and buttons
            document.querySelectorAll('input').forEach(el => {
                el.addEventListener('focus', () => {
                    if (voiceMode && 'speechSynthesis' in window) {
                        const label = el.previousElementSibling ? el.previousElementSibling.textContent.trim() : el.placeholder;
                        speechSynthesis.speak(new SpeechSynthesisUtterance(label));
                    }
                });
            });

            document.querySelectorAll('select').forEach(el => {
                el.addEventListener('focus', () => {
                    if (voiceMode && 'speechSynthesis' in window) {
                        const label = el.previousElementSibling ? el.previousElementSibling.textContent.trim() : '–í—ã–±–æ—Ä';
                        speechSynthesis.speak(new SpeechSynthesisUtterance(label));
                    }
                });
                el.addEventListener('change', () => {
                    if (voiceMode && 'speechSynthesis' in window) {
                        const selected = el.options[el.selectedIndex].text;
                        speechSynthesis.speak(new SpeechSynthesisUtterance('–í—ã–±—Ä–∞–Ω–æ: ' + selected));
                    }
                });
            });

            document.querySelectorAll('button').forEach(el => {
                el.addEventListener('click', () => {
                    if (voiceMode && 'speechSynthesis' in window) {
                        const text = el.textContent.replace(/[^\w\s–∞-—è—ë]/gi, '').trim();
                        speechSynthesis.speak(new SpeechSynthesisUtterance(text));
                    }
                });
            });
        </script>
        </html>
        """)

    @app.route('/api/submit', methods=['POST'])
    def api_submit():
        feature_type = request.form['feature_type']
        description = request.form['description']
        address = request.form['address']
        photo = request.files['photo']
        if photo and photo.filename:
            filename = secure_filename(photo.filename)
            photo_path = filename
            photo.save(photo_path)
        else:
            photo_path = ""
        nav_system.db.add_user_submission(feature_type, description, address, photo_path)
        return redirect(url_for('submit_page'))

    @app.route('/admin')
    def admin_page():
        submissions = nav_system.db.get_pending_submissions()
        return render_template_string("""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>–ê–¥–º–∏–Ω –ø–∞–Ω–µ–ª—å</title>
            <style>
                * { margin: 0; padding: 0; box-sizing: border-box; }
                body {
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    padding: 20px;
                }
                .container {
                    max-width: 1200px;
                    margin: 0 auto;
                    background: white;
                    border-radius: 20px;
                    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                    overflow: hidden;
                }
                .header {
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    padding: 30px;
                    text-align: center;
                }
                .header h1 { font-size: 2.5em; margin-bottom: 10px; }
                .header p { font-size: 1.2em; opacity: 0.9; }
                .content {
                    padding: 30px;
                }
                .submission {
                    border: 2px solid #f0f0f0;
                    border-radius: 10px;
                    padding: 20px;
                    margin: 20px 0;
                    background: #fafafa;
                }
                .submission h3 {
                    margin-bottom: 10px;
                    color: #667eea;
                }
                .submission p {
                    margin: 5px 0;
                }
                .submission img {
                    max-width: 300px;
                    margin: 10px 0;
                    border-radius: 8px;
                }
                .btn {
                    padding: 15px 25px;
                    border: none;
                    border-radius: 12px;
                    font-size: 1.1em;
                    font-weight: 600;
                    cursor: pointer;
                    transition: all 0.3s ease;
                    margin: 8px;
                    min-width: 140px;
                    position: relative;
                    overflow: hidden;
                    text-transform: uppercase;
                    letter-spacing: 0.5px;
                }
                .btn::before {
                    content: '';
                    position: absolute;
                    top: 0;
                    left: -100%;
                    width: 100%;
                    height: 100%;
                    background: linear-gradient(90deg, transparent, rgba(255,255,255,0.2), transparent);
                    transition: left 0.5s;
                }
                .btn:hover::before {
                    left: 100%;
                }
                .btn-approve {
                    background: linear-gradient(135deg, #10b981, #059669);
                    color: white;
                    box-shadow: 0 4px 15px rgba(16, 185, 129, 0.4);
                }
                .btn-approve:hover {
                    background: linear-gradient(135deg, #059669, #047857);
                    transform: translateY(-3px) scale(1.05);
                    box-shadow: 0 8px 25px rgba(16, 185, 129, 0.6);
                }
                .btn-reject {
                    background: linear-gradient(135deg, #ef4444, #dc2626);
                    color: white;
                    box-shadow: 0 4px 15px rgba(239, 68, 68, 0.4);
                }
                .btn-reject:hover {
                    background: linear-gradient(135deg, #dc2626, #b91c1c);
                    transform: translateY(-3px) scale(1.05);
                    box-shadow: 0 8px 25px rgba(239, 68, 68, 0.6);
                }
                .btn-secondary {
                    background: linear-gradient(135deg, #f0f0f0, #e0e0e0);
                    color: #333;
                    box-shadow: 0 4px 15px rgba(0,0,0,0.1);
                }
                .btn-secondary:hover {
                    background: linear-gradient(135deg, #e0e0e0, #d0d0d0);
                    transform: translateY(-2px) scale(1.02);
                    box-shadow: 0 6px 20px rgba(0,0,0,0.15);
                }
                .no-submissions {
                    text-align: center;
                    padding: 50px;
                    color: #666;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>üîß –ê–¥–º–∏–Ω –ø–∞–Ω–µ–ª—å</h1>
                    <p>–£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –æ–±—ä–µ–∫—Ç–∞–º–∏ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏</p>
                    <div class="admin-links">
                        <a href="/admin/districts">–°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ —Ä–∞–π–æ–Ω–∞–º</a>
                        <a href="/admin/change_password">–ò–∑–º–µ–Ω–∏—Ç—å –ø–∞—Ä–æ–ª—å</a>
                        <a href="/admin/add_admin">–î–æ–±–∞–≤–∏—Ç—å –∞–¥–º–∏–Ω–∞</a>
                        <a href="/admin/logout">–í—ã–π—Ç–∏</a>
                    </div>
                </div>
                <div class="content">
                    {% if submissions %}
                        {% for sub in submissions %}
                        <div class="submission" data-id="{{ sub[0] }}">
                            <h3>{{ sub[1].replace('_', ' ').title() }}</h3>
                            <p><strong>–û–ø–∏—Å–∞–Ω–∏–µ:</strong> {{ sub[2] }}</p>
                            <p><strong>–ê–¥—Ä–µ—Å:</strong> {{ sub[3] }}</p>
                            <p><strong>–û—Ç–ø—Ä–∞–≤–∏—Ç–µ–ª—å:</strong> {{ sub[6] or '–ê–Ω–æ–Ω–∏–º' }}</p>
                            {% if sub[4] %}
                            <img src="/uploads/{{ sub[4] }}" alt="–§–æ—Ç–æ –æ–±—ä–µ–∫—Ç–∞">
                            {% endif %}
                            <button class="btn btn-approve" onclick="approve({{ sub[0] }})">‚úÖ –û–¥–æ–±—Ä–∏—Ç—å</button>
                            <button class="btn btn-reject" onclick="reject({{ sub[0] }})">‚ùå –û—Ç–∫–ª–æ–Ω–∏—Ç—å</button>
                        </div>
                        {% endfor %}
                    {% else %}
                        <div class="no-submissions">
                            <h2>–ù–µ—Ç –æ–∂–∏–¥–∞—é—â–∏—Ö –ø–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–π</h2>
                            <p>–í—Å–µ –æ–±—ä–µ–∫—Ç—ã –ø—Ä–æ–≤–µ—Ä–µ–Ω—ã</p>
                        </div>
                    {% endif %}
                    <a href="/" class="btn btn-secondary">–ù–∞–∑–∞–¥ –∫ –Ω–∞–≤–∏–≥–∞—Ü–∏–∏</a>
                </div>
            </div>
            <script>
                function approve(id) {
                    fetch('/api/approve/' + id, { method: 'POST' })
                        .then(response => {
                            if (response.ok) {
                                removeSubmission(id);
                            } else {
                                alert('–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–¥–æ–±—Ä–µ–Ω–∏–∏');
                            }
                        });
                }
                function reject(id) {
                    if (confirm('–û—Ç–∫–ª–æ–Ω–∏—Ç—å –æ–±—ä–µ–∫—Ç?')) {
                        fetch('/api/reject/' + id, { method: 'POST' })
                            .then(response => {
                                if (response.ok) {
                                    removeSubmission(id);
                                } else {
                                    alert('–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ—Ç–∫–ª–æ–Ω–µ–Ω–∏–∏');
                                }
                            });
                        }
                    }
                function removeSubmission(id) {
                    const submission = document.querySelector(`[data-id="${id}"]`);
                    if (submission) {
                        submission.remove();
                        // Check if no submissions left
                        const submissions = document.querySelectorAll('.submission');
                        if (submissions.length === 0) {
                            document.querySelector('.content').innerHTML = `
                                <div class="no-submissions">
                                    <h2>–ù–µ—Ç –æ–∂–∏–¥–∞—é—â–∏—Ö –ø–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–π</h2>
                                    <p>–í—Å–µ –æ–±—ä–µ–∫—Ç—ã –ø—Ä–æ–≤–µ—Ä–µ–Ω—ã</p>
                                </div>
                                <a href="/" class="btn btn-secondary">–ù–∞–∑–∞–¥ –∫ –Ω–∞–≤–∏–≥–∞—Ü–∏–∏</a>
                            `;
                        }
                    }
                }
            </script>
        </body>
        </html>
        """, submissions=submissions)

    @app.route('/api/approve/<int:submission_id>', methods=['POST'])
    def api_approve(submission_id):
        nav_system.db.approve_submission(submission_id)
        return '', 200

    @app.route('/api/reject/<int:submission_id>', methods=['POST'])
    def api_reject(submission_id):
        try:
            conn = sqlite3.connect(nav_system.db.db_path, timeout=10)
            cursor = conn.cursor()
            cursor.execute("UPDATE user_submissions SET status = 'rejected' WHERE id = ?", (submission_id,))
            conn.commit()
        except sqlite3.OperationalError as e:
            return jsonify({"error": "Database locked, try again"}), 500
        finally:
            if 'conn' in locals():
                conn.close()
        return '', 200

    @app.before_request
    def require_admin():
        if request.path.startswith('/admin') and request.path != '/admin/login' and request.path != '/admin/change_password':
            if not session.get('admin'):
                return redirect(url_for('admin_login'))

    @app.route('/admin/login', methods=['GET', 'POST'])
    def admin_login():
        if request.method == 'POST':
            username = request.form['username']
            password = request.form['password']
            conn = sqlite3.connect(nav_system.db.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT password, must_change_password FROM admins WHERE username = ?", (username,))
            row = cursor.fetchone()
            conn.close()
            if row and check_password_hash(row[0], password):
                session['admin'] = username
                if row[1]:
                    return redirect(url_for('change_password'))
                return redirect(url_for('admin_page'))
            flash('–ù–µ–≤–µ—Ä–Ω—ã–µ —É—á–µ—Ç–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ')
        return render_template_string("""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>–í—Ö–æ–¥ –≤ –∞–¥–º–∏–Ω –ø–∞–Ω–µ–ª—å</title>
            <style>
                * { margin: 0; padding: 0; box-sizing: border-box; }
                body {
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                }
                .container {
                    background: white;
                    border-radius: 20px;
                    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                    padding: 40px;
                    width: 100%;
                    max-width: 400px;
                }
                .form-group {
                    margin-bottom: 20px;
                }
                .form-group label {
                    display: block;
                    margin-bottom: 8px;
                    font-weight: 600;
                    color: #333;
                }
                .form-group input {
                    width: 100%;
                    padding: 12px;
                    border: 2px solid #e0e0e0;
                    border-radius: 8px;
                    font-size: 1em;
                    transition: border-color 0.3s;
                }
                .form-group input:focus {
                    outline: none;
                    border-color: #667eea;
                }
                .btn {
                    width: 100%;
                    padding: 15px;
                    border: none;
                    border-radius: 8px;
                    font-size: 1.1em;
                    font-weight: 600;
                    cursor: pointer;
                    transition: all 0.3s;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                }
                .btn:hover {
                    transform: translateY(-2px);
                    box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
                }
                .flash {
                    color: red;
                    margin-bottom: 20px;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <h1 style="text-align: center; margin-bottom: 30px;">–í—Ö–æ–¥ –≤ –∞–¥–º–∏–Ω –ø–∞–Ω–µ–ª—å</h1>
                {% with messages = get_flashed_messages() %}
                    {% if messages %}
                        <div class="flash">{{ messages[0] }}</div>
                    {% endif %}
                {% endwith %}
                <form method="post">
                    <div class="form-group">
                        <label for="username">–ò–º—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è:</label>
                        <input type="text" id="username" name="username" required>
                    </div>
                    <div class="form-group">
                        <label for="password">–ü–∞—Ä–æ–ª—å:</label>
                        <input type="password" id="password" name="password" required>
                    </div>
                    <button type="submit" class="btn">–í–æ–π—Ç–∏</button>
                </form>
            </div>
        </body>
        </html>
        """)

    @app.route('/admin/change_password', methods=['GET', 'POST'])
    def change_password():
        if not session.get('admin'):
            return redirect(url_for('admin_login'))
        if request.method == 'POST':
            new_password = request.form['new_password']
            confirm_password = request.form['confirm_password']
            if new_password != confirm_password:
                flash('–ü–∞—Ä–æ–ª–∏ –Ω–µ —Å–æ–≤–ø–∞–¥–∞—é—Ç')
                return redirect(request.url)
            conn = sqlite3.connect(nav_system.db.db_path)
            cursor = conn.cursor()
            cursor.execute("UPDATE admins SET password = ?, must_change_password = 0 WHERE username = ?",
                           (generate_password_hash(new_password), session['admin']))
            conn.commit()
            conn.close()
            flash('–ü–∞—Ä–æ–ª—å –∏–∑–º–µ–Ω–µ–Ω')
            return redirect(url_for('admin_page'))
        return render_template_string("""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>–ò–∑–º–µ–Ω–∏—Ç—å –ø–∞—Ä–æ–ª—å</title>
            <style>
                * { margin: 0; padding: 0; box-sizing: border-box; }
                body {
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                }
                .container {
                    background: white;
                    border-radius: 20px;
                    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                    padding: 40px;
                    width: 100%;
                    max-width: 400px;
                }
                .form-group {
                    margin-bottom: 20px;
                }
                .form-group label {
                    display: block;
                    margin-bottom: 8px;
                    font-weight: 600;
                    color: #333;
                }
                .form-group input {
                    width: 100%;
                    padding: 12px;
                    border: 2px solid #e0e0e0;
                    border-radius: 8px;
                    font-size: 1em;
                    transition: border-color 0.3s;
                }
                .form-group input:focus {
                    outline: none;
                    border-color: #667eea;
                }
                .btn {
                    width: 100%;
                    padding: 15px;
                    border: none;
                    border-radius: 8px;
                    font-size: 1.1em;
                    font-weight: 600;
                    cursor: pointer;
                    transition: all 0.3s;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                }
                .btn:hover {
                    transform: translateY(-2px);
                    box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
                }
                .flash {
                    color: red;
                    margin-bottom: 20px;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <h1 style="text-align: center; margin-bottom: 30px;">–ò–∑–º–µ–Ω–∏—Ç—å –ø–∞—Ä–æ–ª—å</h1>
                {% with messages = get_flashed_messages() %}
                    {% if messages %}
                        <div class="flash">{{ messages[0] }}</div>
                    {% endif %}
                {% endwith %}
                <form method="post">
                    <div class="form-group">
                        <label for="new_password">–ù–æ–≤—ã–π –ø–∞—Ä–æ–ª—å:</label>
                        <input type="password" id="new_password" name="new_password" required>
                    </div>
                    <div class="form-group">
                        <label for="confirm_password">–ü–æ–¥—Ç–≤–µ—Ä–¥–∏—Ç—å –ø–∞—Ä–æ–ª—å:</label>
                        <input type="password" id="confirm_password" name="confirm_password" required>
                    </div>
                    <button type="submit" class="btn">–ò–∑–º–µ–Ω–∏—Ç—å</button>
                </form>
            </div>
        </body>
        </html>
        """)

    @app.route('/admin/add_admin', methods=['GET', 'POST'])
    def add_admin():
        if not session.get('admin'):
            return redirect(url_for('admin_login'))
        if request.method == 'POST':
            username = request.form['username']
            password = request.form['password']
            conn = sqlite3.connect(nav_system.db.db_path)
            cursor = conn.cursor()
            try:
                cursor.execute("INSERT INTO admins (username, password, must_change_password) VALUES (?, ?, ?)",
                               (username, generate_password_hash(password), 0))
                conn.commit()
                flash('–ê–¥–º–∏–Ω –¥–æ–±–∞–≤–ª–µ–Ω')
            except sqlite3.IntegrityError:
                flash('–ò–º—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç')
            conn.close()
            return redirect(url_for('admin_page'))
        return render_template_string("""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>–î–æ–±–∞–≤–∏—Ç—å –∞–¥–º–∏–Ω–∞</title>
            <style>
                * { margin: 0; padding: 0; box-sizing: border-box; }
                body {
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                }
                .container {
                    background: white;
                    border-radius: 20px;
                    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                    padding: 40px;
                    width: 100%;
                    max-width: 400px;
                }
                .form-group {
                    margin-bottom: 20px;
                }
                .form-group label {
                    display: block;
                    margin-bottom: 8px;
                    font-weight: 600;
                    color: #333;
                }
                .form-group input {
                    width: 100%;
                    padding: 12px;
                    border: 2px solid #e0e0e0;
                    border-radius: 8px;
                    font-size: 1em;
                    transition: border-color 0.3s;
                }
                .form-group input:focus {
                    outline: none;
                    border-color: #667eea;
                }
                .btn {
                    width: 100%;
                    padding: 15px;
                    border: none;
                    border-radius: 8px;
                    font-size: 1.1em;
                    font-weight: 600;
                    cursor: pointer;
                    transition: all 0.3s;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                }
                .btn:hover {
                    transform: translateY(-2px);
                    box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
                }
                .flash {
                    color: red;
                    margin-bottom: 20px;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <h1 style="text-align: center; margin-bottom: 30px;">–î–æ–±–∞–≤–∏—Ç—å –∞–¥–º–∏–Ω–∞</h1>
                {% with messages = get_flashed_messages() %}
                    {% if messages %}
                        <div class="flash">{{ messages[0] }}</div>
                    {% endif %}
                {% endwith %}
                <form method="post">
                    <div class="form-group">
                        <label for="username">–ò–º—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è:</label>
                        <input type="text" id="username" name="username" required>
                    </div>
                    <div class="form-group">
                        <label for="password">–ü–∞—Ä–æ–ª—å:</label>
                        <input type="password" id="password" name="password" required>
                    </div>
                    <button type="submit" class="btn">–î–æ–±–∞–≤–∏—Ç—å</button>
                </form>
            </div>
        </body>
        </html>
        """)

    @app.route('/admin/logout')
    def logout():
        session.pop('admin', None)
        return redirect(url_for('admin_login'))

    @app.route('/admin/districts')
    def admin_districts():
        if not session.get('admin'):
            return redirect(url_for('admin_login'))
        stats = get_district_statistics(nav_system.db.db_path)
        return render_template_string("""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>–°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ —Ä–∞–π–æ–Ω–∞–º</title>
            <link href='https://unpkg.com/maplibre-gl@3.6.2/dist/maplibre-gl.css' rel='stylesheet' />
            <script src='https://unpkg.com/maplibre-gl@3.6.2/dist/maplibre-gl.js'></script>
            <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
            <style>
                * { margin: 0; padding: 0; box-sizing: border-box; }
                body {
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    padding: 20px;
                }
                .container {
                    max-width: 1400px;
                    margin: 0 auto;
                    background: white;
                    border-radius: 20px;
                    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                    overflow: hidden;
                }
                .header {
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    padding: 30px;
                    text-align: center;
                }
                .header h1 { font-size: 2.5em; margin-bottom: 10px; }
                .header p { font-size: 1.2em; opacity: 0.9; }
                .header .links { margin-top: 20px; }
                .header .links a { color: white; margin: 0 10px; text-decoration: none; }
                .content {
                    padding: 30px;
                }
                .stats-grid {
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
                    gap: 20px;
                    margin-bottom: 30px;
                }
                .stat-card {
                    background: #f8f9fa;
                    border-radius: 10px;
                    padding: 20px;
                    border-left: 4px solid #667eea;
                }
                .stat-card h3 {
                    color: #667eea;
                    margin-bottom: 15px;
                }
                .stat-value {
                    font-size: 2em;
                    font-weight: bold;
                    color: #333;
                }
                .charts-container {
                    display: grid;
                    grid-template-columns: 1fr 1fr;
                    gap: 30px;
                    margin-bottom: 30px;
                }
                .chart-wrapper {
                    background: white;
                    border-radius: 10px;
                    padding: 20px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                }
                #map {
                    height: 500px;
                    border-radius: 10px;
                    margin-bottom: 30px;
                }
                .table-container {
                    background: white;
                    border-radius: 10px;
                    padding: 20px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                    margin-bottom: 30px;
                }
                table {
                    width: 100%;
                    border-collapse: collapse;
                }
                th, td {
                    padding: 12px;
                    text-align: left;
                    border-bottom: 1px solid #ddd;
                }
                th {
                    background: #f8f9fa;
                    font-weight: 600;
                }
                .btn {
                    padding: 12px 20px;
                    border: none;
                    border-radius: 8px;
                    font-size: 1em;
                    font-weight: 600;
                    cursor: pointer;
                    transition: all 0.3s;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    text-decoration: none;
                    display: inline-block;
                }
                .btn:hover {
                    transform: translateY(-2px);
                    box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
                }
                .export-btn {
                    margin-bottom: 20px;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>üìä –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ –ø–æ —Ä–∞–π–æ–Ω–∞–º –¢—É–ª—ã</h1>
                    <p>–ê–Ω–∞–ª–∏–∑ –æ–±—ä–µ–∫—Ç–æ–≤ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ –≤ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–∏–≤–Ω—ã—Ö —Ä–∞–π–æ–Ω–∞—Ö</p>
                    <div class="links">
                        <a href="/admin">‚Üê –ê–¥–º–∏–Ω –ø–∞–Ω–µ–ª—å</a>
                        <a href="/admin/logout">–í—ã–π—Ç–∏</a>
                    </div>
                </div>
                <div class="content">
                    <div class="stats-grid">
                        {% for district, data in stats.items() %}
                        <div class="stat-card">
                            <h3>{{ data.name }}</h3>
                            <div class="stat-value">{{ data.total_objects }}</div>
                            <p>–æ–±—ä–µ–∫—Ç–æ–≤ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏</p>
                        </div>
                        {% endfor %}
                    </div>

                    <div class="charts-container">
                        <div class="chart-wrapper">
                            <canvas id="mobilityChart"></canvas>
                        </div>
                        <div class="chart-wrapper">
                            <canvas id="districtChart"></canvas>
                        </div>
                    </div>

                    <div id="map"></div>


                    <div class="table-container">
                        <h3>–î–µ—Ç–∞–ª—å–Ω–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞</h3>
                        <table>
                            <thead>
                                <tr>
                                    <th>–†–∞–π–æ–Ω</th>
                                    <th>–í—Å–µ–≥–æ –æ–±—ä–µ–∫—Ç–æ–≤</th>
                                    <th>–ö–æ–ª—è—Å–æ—á–Ω–∏–∫–∏</th>
                                    <th>–°–ª–∞–±–æ–≤–∏–¥—è—â–∏–µ</th>
                                    <th>–û–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å</th>
                                    <th>–î—Ä—É–≥–∏–µ</th>
                                </tr>
                            </thead>
                            <tbody>
                                {% for district, data in stats.items() %}
                                <tr>
                                    <td>{{ data.name }}</td>
                                    <td>{{ data.total_objects }}</td>
                                    <td>{{ data.by_mobility['–∫–æ–ª—è—Å–æ—á–Ω–∏–∫'] }}</td>
                                    <td>{{ data.by_mobility['—Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–π'] }}</td>
                                    <td>{{ data.by_mobility['–æ–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å'] }}</td>
                                    <td>{{ data.by_mobility['–¥—Ä—É–≥–∏–µ'] }}</td>
                                </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    </div>

                    <a href="/admin/export_districts" class="btn export-btn">üì• –≠–∫—Å–ø–æ—Ä—Ç –≤ Excel</a>
                    <a href="/admin" class="btn">‚Üê –ù–∞–∑–∞–¥</a>
                </div>
            </div>

            <script>
                // –î–∞–Ω–Ω—ã–µ –¥–ª—è –≥—Ä–∞—Ñ–∏–∫–æ–≤
                const stats = {{ stats|tojson }};
                const districts = Object.keys(stats);
                const totalObjects = districts.map(d => stats[d].total_objects);
                const wheelchair = districts.map(d => stats[d].by_mobility['–∫–æ–ª—è—Å–æ—á–Ω–∏–∫']);
                const visuallyImpaired = districts.map(d => stats[d].by_mobility['—Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–π']);
                const cane = districts.map(d => stats[d].by_mobility['–æ–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å']);
                const other = districts.map(d => stats[d].by_mobility['–¥—Ä—É–≥–∏–µ']);

                // –ì—Ä–∞—Ñ–∏–∫ –ø–æ —Ä–∞–π–æ–Ω–∞–º
                const ctx1 = document.getElementById('districtChart').getContext('2d');
                new Chart(ctx1, {
                    type: 'bar',
                    data: {
                        labels: districts.map(d => stats[d].name),
                        datasets: [{
                            label: '–í—Å–µ–≥–æ –æ–±—ä–µ–∫—Ç–æ–≤',
                            data: totalObjects,
                            backgroundColor: '#667eea',
                            borderColor: '#667eea',
                            borderWidth: 1
                        }]
                    },
                    options: {
                        responsive: true,
                        plugins: {
                            title: {
                                display: true,
                                text: '–û–±—ä–µ–∫—Ç—ã –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ –ø–æ —Ä–∞–π–æ–Ω–∞–º'
                            }
                        },
                        scales: {
                            y: {
                                beginAtZero: true
                            }
                        }
                    }
                });

                // –ì—Ä–∞—Ñ–∏–∫ –ø–æ —Ç–∏–ø–∞–º –º–æ–±–∏–ª—å–Ω–æ—Å—Ç–∏
                const ctx2 = document.getElementById('mobilityChart').getContext('2d');
                new Chart(ctx2, {
                    type: 'radar',
                    data: {
                        labels: ['–ö–æ–ª—è—Å–æ—á–Ω–∏–∫–∏', '–°–ª–∞–±–æ–≤–∏–¥—è—â–∏–µ', '–û–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å', '–î—Ä—É–≥–∏–µ'],
                        datasets: districts.map((district, i) => ({
                            label: stats[district].name,
                            data: [
                                stats[district].by_mobility['–∫–æ–ª—è—Å–æ—á–Ω–∏–∫'],
                                stats[district].by_mobility['—Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–π'],
                                stats[district].by_mobility['–æ–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å'],
                                stats[district].by_mobility['–¥—Ä—É–≥–∏–µ']
                            ],
                            backgroundColor: `rgba(${50 + i*50}, ${100 + i*30}, ${200 - i*40}, 0.2)`,
                            borderColor: `rgba(${50 + i*50}, ${100 + i*30}, ${200 - i*40}, 1)`,
                            borderWidth: 2
                        }))
                    },
                    options: {
                        responsive: true,
                        plugins: {
                            title: {
                                display: true,
                                text: '–†–∞—Å–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ –ø–æ —Ç–∏–ø–∞–º –∏–Ω–≤–∞–ª–∏–¥–Ω–æ—Å—Ç–∏'
                            }
                        }
                    }
                });

                // –ö–∞—Ä—Ç–∞ —Å —Ä–∞–π–æ–Ω–∞–º–∏
                const map = new maplibregl.Map({
                    container: 'map',
                    style: 'https://basemaps.cartocdn.com/gl/voyager-gl-style/style.json',
                    center: [37.6175, 54.1931],
                    zoom: 13,
                    pitch: 30,
                    bearing: 0
                });
                map.addControl(new maplibregl.NavigationControl());
                map.addControl(new maplibregl.GeolocateControl({
                    positionOptions: { enableHighAccuracy: true },
                    trackUserLocation: true
                }));

                // –î–æ–±–∞–≤–ª—è–µ–º —Ä–∞–π–æ–Ω—ã –∫–∞–∫ –ø—Ä—è–º–æ—É–≥–æ–ª—å–Ω–∏–∫–∏
                const districtBounds = {
                    '–¶–µ–Ω—Ç—Ä–∞–ª—å–Ω—ã–π': [[54.185, 37.605], [54.200, 37.630]],
                    '–°–æ–≤–µ—Ç—Å–∫–∏–π': [[54.200, 37.600], [54.220, 37.640]],
                    '–ü—Ä–∏–≤–æ–∫–∑–∞–ª—å–Ω—ã–π': [[54.185, 37.630], [54.205, 37.650]],
                    '–ó–∞—Ä–µ—á–µ–Ω—Å–∫–∏–π': [[54.175, 37.590], [54.195, 37.620]],
                    '–ü—Ä–æ–ª–µ—Ç–∞—Ä—Å–∫–∏–π': [[54.180, 37.580], [54.200, 37.605]]
                };

                const colors = {
                    '–¶–µ–Ω—Ç—Ä–∞–ª—å–Ω—ã–π': '#ff6b6b',
                    '–°–æ–≤–µ—Ç—Å–∫–∏–π': '#4ecdc4',
                    '–ü—Ä–∏–≤–æ–∫–∑–∞–ª—å–Ω—ã–π': '#45b7d1',
                    '–ó–∞—Ä–µ—á–µ–Ω—Å–∫–∏–π': '#f9ca24',
                    '–ü—Ä–æ–ª–µ—Ç–∞—Ä—Å–∫–∏–π': '#6c5ce7'
                };

                // –î–∞–Ω–Ω—ã–µ –ø–æ–ª–∏–≥–æ–Ω–æ–≤ —Ä–∞–π–æ–Ω–æ–≤ [lat, lon] - –Ω–µ –ø–µ—Ä–µ—Å–µ–∫–∞—é—â–∏–µ—Å—è –≥—Ä–∞–Ω–∏—Ü—ã
                const districtPolygons = {
                    '–¶–µ–Ω—Ç—Ä–∞–ª—å–Ω—ã–π': [[54.180, 37.600], [54.180, 37.610], [54.180, 37.620], [54.180, 37.625], [54.190, 37.625], [54.200, 37.625], [54.200, 37.615], [54.200, 37.605], [54.200, 37.600], [54.190, 37.600]],
                    '–°–æ–≤–µ—Ç—Å–∫–∏–π': [[54.200, 37.600], [54.200, 37.610], [54.200, 37.620], [54.200, 37.630], [54.200, 37.640], [54.210, 37.640], [54.220, 37.640], [54.220, 37.630], [54.220, 37.620], [54.220, 37.610], [54.220, 37.600], [54.210, 37.600]],
                    '–ü—Ä–∏–≤–æ–∫–∑–∞–ª—å–Ω—ã–π': [[54.180, 37.625], [54.180, 37.635], [54.180, 37.645], [54.180, 37.650], [54.190, 37.650], [54.200, 37.650], [54.200, 37.640], [54.200, 37.630], [54.200, 37.625], [54.190, 37.625]],
                    '–ó–∞—Ä–µ—á–µ–Ω—Å–∫–∏–π': [[54.170, 37.580], [54.170, 37.590], [54.170, 37.600], [54.170, 37.610], [54.175, 37.610], [54.180, 37.610], [54.185, 37.610], [54.185, 37.600], [54.185, 37.590], [54.185, 37.580], [54.180, 37.580], [54.175, 37.580]],
                    '–ü—Ä–æ–ª–µ—Ç–∞—Ä—Å–∫–∏–π': [[54.185, 37.580], [54.185, 37.590], [54.185, 37.600], [54.185, 37.610], [54.190, 37.610], [54.195, 37.610], [54.200, 37.610], [54.200, 37.600], [54.200, 37.590], [54.200, 37.580], [54.195, 37.580], [54.190, 37.580]]
                };

    
                // Modal close
                const modal = document.getElementById('videoModal');
                const closeBtn = document.getElementsByClassName('close')[0];
                closeBtn.onclick = function() {
                    modal.style.display = 'none';
                    const video = document.getElementById('samovarVideo');
                    video.pause();
                    const audio = document.getElementById('bgMusic');
                    audio.pause();
                };
                window.onclick = function(event) {
                    if (event.target == modal) {
                        modal.style.display = 'none';
                        const video = document.getElementById('samovarVideo');
                        video.pause();
                        const audio = document.getElementById('bgMusic');
                        audio.pause();
                    }
                };
    
                // Randomize PNG positions
                pngElements.forEach(id => {
                    const img = document.getElementById(id);
                    const randomLeft = Math.random() * 80 + 10; // 10% to 90%
                    img.style.top = 'calc(100vh - 150px)';
                    img.style.left = randomLeft + '%';
                });
    
                map.on('load', () => {
                    console.log("MapLibre –≥–æ—Ç–æ–≤–∞ ‚Äî –≤—Å—ë –∏–¥–µ–∞–ª—å–Ω–æ!");

                    // Try to load real districts from geojson
                    fetch('/tula_districts/tula_administrative_districts.geojson')
                    .then(response => {
                        if (!response.ok) throw new Error();
                        return response.json();
                    })
                    .then(data => {
                        // Use real data
                        data.features.forEach(feature => {
                            const name = feature.properties.name;
                            const key = name.split(' ')[0]; // e.g. –¶–µ–Ω—Ç—Ä–∞–ª—å–Ω—ã–π
                            const color = colors[key] || '#888';
                            map.addSource(name, {
                                type: 'geojson',
                                data: feature
                            });
                            map.addLayer({
                                id: name,
                                type: 'fill',
                                source: name,
                                paint: { 'fill-color': color, 'fill-opacity': 0.5 }
                            });
                            map.addLayer({
                                id: name + '-border',
                                type: 'line',
                                source: name,
                                paint: { 'line-color': color, 'line-width': 3 }
                            });
                            map.on('click', name + '-border', (e) => {
                                const districtKey = Object.keys(stats).find(k => stats[k].name === name);
                                if (districtKey) {
                                    const d = stats[districtKey];
                                    new maplibregl.Popup()
                                        .setLngLat(e.lngLat)
                                        .setHTML(`
                                            <b>${d.name}</b><br>
                                            –û–±—ä–µ–∫—Ç–æ–≤: ${d.total_objects}<br>
                                            –ö–æ–ª—è—Å–æ—á–Ω–∏–∫–∏: ${d.by_mobility['–∫–æ–ª—è—Å–æ—á–Ω–∏–∫']}<br>
                                            –°–ª–∞–±–æ–≤–∏–¥—è—â–∏–µ: ${d.by_mobility['—Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–π']}<br>
                                            –û–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å: ${d.by_mobility['–æ–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å']}
                                        `)
                                        .addTo(map);
                                }
                            });
                            map.on('mouseenter', name + '-border', () => {
                                map.getCanvas().style.cursor = 'pointer';
                            });
                            map.on('mouseleave', name + '-border', () => {
                                map.getCanvas().style.cursor = '';
                            });
                        });
                    })
                    .catch(() => {
                        // Fallback to hardcoded
                        const districtPolygons = {
                            '–¶–µ–Ω—Ç—Ä–∞–ª—å–Ω—ã–π': [[54.180, 37.600], [54.180, 37.610], [54.180, 37.620], [54.180, 37.625], [54.190, 37.625], [54.200, 37.625], [54.200, 37.615], [54.200, 37.605], [54.200, 37.600], [54.190, 37.600]],
                            '–°–æ–≤–µ—Ç—Å–∫–∏–π': [[54.200, 37.600], [54.200, 37.610], [54.200, 37.620], [54.200, 37.630], [54.200, 37.640], [54.210, 37.640], [54.220, 37.640], [54.220, 37.630], [54.220, 37.620], [54.220, 37.610], [54.220, 37.600], [54.210, 37.600]],
                            '–ü—Ä–∏–≤–æ–∫–∑–∞–ª—å–Ω—ã–π': [[54.180, 37.625], [54.180, 37.635], [54.180, 37.645], [54.180, 37.650], [54.190, 37.650], [54.200, 37.650], [54.200, 37.640], [54.200, 37.630], [54.200, 37.625], [54.190, 37.625]],
                            '–ó–∞—Ä–µ—á–µ–Ω—Å–∫–∏–π': [[54.170, 37.580], [54.170, 37.590], [54.170, 37.600], [54.170, 37.610], [54.175, 37.610], [54.180, 37.610], [54.185, 37.610], [54.185, 37.600], [54.185, 37.590], [54.185, 37.580], [54.180, 37.580], [54.175, 37.580]],
                            '–ü—Ä–æ–ª–µ—Ç–∞—Ä—Å–∫–∏–π': [[54.185, 37.580], [54.185, 37.590], [54.185, 37.600], [54.185, 37.610], [54.190, 37.610], [54.195, 37.610], [54.200, 37.610], [54.200, 37.600], [54.200, 37.590], [54.200, 37.580], [54.195, 37.580], [54.190, 37.580]]
                        };
                        Object.keys(districtPolygons).forEach(district => {
                            const polygon = districtPolygons[district];
                            const coords = polygon.map(p => [p[1], p[0]]);
                            coords.push(coords[0]);
                            map.addSource(district, {
                                type: 'geojson',
                                data: { type: 'Feature', geometry: { type: 'Polygon', coordinates: [coords] } }
                            });
                            map.addLayer({
                                id: district,
                                type: 'fill',
                                source: district,
                                paint: { 'fill-color': colors[district], 'fill-opacity': 0.5 }
                            });
                            map.addLayer({
                                id: district + '-border',
                                type: 'line',
                                source: district,
                                paint: { 'line-color': colors[district], 'line-width': 3 }
                            });
                            map.on('click', district + '-border', (e) => {
                                new maplibregl.Popup()
                                    .setLngLat(e.lngLat)
                                    .setHTML(`
                                        <b>${stats[district].name}</b><br>
                                        –û–±—ä–µ–∫—Ç–æ–≤: ${stats[district].total_objects}<br>
                                        –ö–æ–ª—è—Å–æ—á–Ω–∏–∫–∏: ${stats[district].by_mobility['–∫–æ–ª—è—Å–æ—á–Ω–∏–∫']}<br>
                                        –°–ª–∞–±–æ–≤–∏–¥—è—â–∏–µ: ${d.by_mobility['—Å–ª–∞–±–æ–≤–∏–¥—è—â–∏–π']}<br>
                                        –û–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å: ${d.by_mobility['–æ–ø–æ—Ä–∞ –Ω–∞ —Ç—Ä–æ—Å—Ç—å']}
                                    `)
                                    .addTo(map);
                            });
                            map.on('mouseenter', district + '-border', () => {
                                map.getCanvas().style.cursor = 'pointer';
                            });
                            map.on('mouseleave', district + '-border', () => {
                                map.getCanvas().style.cursor = '';
                            });
                        });
                    });
                });
            </script>
        </body>
        </html>
        """, stats=stats)

    @app.route('/admin/export_districts')
    def export_districts():
        if not session.get('admin'):
            return redirect(url_for('admin_login'))
        stats = get_district_statistics(nav_system.db.db_path)
        filename = export_district_stats_to_excel(stats)
        return send_from_directory('.', filename, as_attachment=True)

    @app.route('/uploads/<filename>')
    def uploaded_file(filename):
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

    @app.route('/tula_districts/<filename>')
    def serve_districts(filename):
        return send_from_directory('tula_districts', filename)

    @app.route('/music/<filename>')
    def serve_music(filename):
        return send_from_directory('music', filename)

    if __name__ == '__main__':
        print("–ó–∞–ø—É—Å–∫ –¥–æ—Å—Ç—É–ø–Ω–æ–π –Ω–∞–≤–∏–≥–∞—Ü–∏–∏...")
        print("–û—Ç–∫—Ä–æ–π—Ç–µ –≤ –±—Ä–∞—É–∑–µ—Ä–µ: http://127.0.0.1:5001")
        app.run(debug=True, host='0.0.0.0', port=5000)

except ImportError:
    print("–î–ª—è –∑–∞–ø—É—Å–∫–∞ –≤–µ–±-–∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å–∞ —É—Å—Ç–∞–Ω–æ–≤–∏—Ç–µ: pip install flask flask-cors requests")
    print("–ò–ª–∏ –∑–∞–ø—É—Å—Ç–∏—Ç–µ —Ç–æ–ª—å–∫–æ –∫–∞–∫ –±–∏–±–ª–∏–æ—Ç–µ–∫—É –±–µ–∑ –≤–µ–±-—Å–µ—Ä–≤–µ—Ä–∞.")
